import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import numpy as np
import io
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# PAGE CONFIG
# ============================================================================
st.set_page_config(
    page_title="Water Consumption Dashboard",
    page_icon="💧",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================================
# CUSTOM CSS
# ============================================================================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .main { background-color: #f8fafc; }
    h1, h2, h3 { color: #1e293b; font-weight: 700; }

    .stTabs [data-baseweb="tab-list"] {
        gap: 0.5rem; background-color: white;
        border-radius: 0.75rem; padding: 0.5rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    }
    .stTabs [data-baseweb="tab"] {
        background-color: transparent; border-radius: 0.5rem;
        padding: 0.75rem 1.5rem; font-weight: 600; color: #64748b; border: none;
    }
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        background-color: #0ea5e9; color: white;
    }
    [data-testid="metric-container"] {
        background: white; border-radius: 12px; padding: 1rem;
        box-shadow: 0 1px 4px rgba(0,0,0,0.08); border-left: 4px solid #0ea5e9;
    }
    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }
    .stSelectbox > div > div { background-color: white; border-radius: 0.5rem; }
    .stMultiSelect > div > div { background-color: white; border-radius: 0.5rem; }
    [data-testid="stSidebar"] { background: white; border-right: 1px solid #e5e7eb; }
    [data-testid="stSidebar"] .stMarkdown h3 { color: #0c4a6e; }
    [data-testid="stFileUploadDropzone"] {
        background: #f0f9ff; border: 2px dashed #0ea5e9; border-radius: 12px;
    }
    .stAlert { border-radius: 10px; }
</style>
""", unsafe_allow_html=True)

# ============================================================================
# CONSTANTS
# ============================================================================
DEFAULT_PRODS = ['PROD 1.3', 'PROD 1.4', 'PROD 1.5', 'PROD 1.6',
                 'PROD 2.5', 'PROD 2.6', 'PROD 2.7']

_COLOR_PALETTE = [
    '#0ea5e9', '#10b981', '#f59e0b', '#ef4444',
    '#8b5cf6', '#ec4899', '#06b6d4', '#f97316',
    '#84cc16', '#6366f1', '#14b8a6', '#e11d48',
    '#a855f7', '#0284c7', '#059669', '#d97706',
]
_UNIT_COLOR_CACHE: dict = {}

def get_unit_color(unit: str) -> str:
    _LEGACY = {
        'PROD 1.3': '#0ea5e9', 'PROD 1.4': '#10b981',
        'PROD 1.5': '#f59e0b', 'PROD 1.6': '#ef4444',
        'PROD 2.5': '#8b5cf6', 'PROD 2.6': '#ec4899',
        'PROD 2.7': '#06b6d4',
    }
    if unit in _LEGACY:
        return _LEGACY[unit]
    if unit not in _UNIT_COLOR_CACHE:
        idx = len(_UNIT_COLOR_CACHE) % len(_COLOR_PALETTE)
        _UNIT_COLOR_CACHE[unit] = _COLOR_PALETTE[idx]
    return _UNIT_COLOR_CACHE[unit]

# ============================================================================
# DATA LOADING
# ============================================================================

def fix_excel_bytes(raw_bytes: bytes) -> bytes:
    EOCD = b'PK\x05\x06'
    idx = raw_bytes.rfind(EOCD)
    if idx != -1:
        return raw_bytes[:idx + 22]
    return raw_bytes


@st.cache_data(show_spinner=False)
def load_raw_data(file_bytes: bytes) -> pd.DataFrame:
    fixed = fix_excel_bytes(file_bytes)
    buf   = io.BytesIO(fixed)

    probe = pd.read_excel(buf, sheet_name=0, header=None, nrows=10)
    header_row = 0
    for i, row in probe.iterrows():
        row_str = ' '.join([str(v).lower() for v in row.values])
        if 'date' in row_str and 'location' in row_str:
            header_row = i
            break

    buf.seek(0)
    df = pd.read_excel(io.BytesIO(fixed), sheet_name=0, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]

    col_map = {}
    for c in df.columns:
        cl = c.lower()
        if 'date' in cl:
            col_map[c] = 'Date'
        elif 'location' in cl:
            col_map[c] = 'Location'
        elif 'pompa' in cl:
            col_map[c] = 'Pompa'
        elif 'water' in cl or 'indicator' in cl:
            col_map[c] = 'Water_Indicator'
    df = df.rename(columns=col_map)

    required = ['Date', 'Location', 'Pompa', 'Water_Indicator']
    for r in required:
        if r not in df.columns:
            st.error(
                f"Column '{r}' not found. "
                f"Please ensure the file contains: Date, Location, Pompa, Water Indicator"
            )
            return pd.DataFrame()

    df['Date']            = pd.to_datetime(df['Date'], errors='coerce')
    df['Water_Indicator'] = pd.to_numeric(df['Water_Indicator'], errors='coerce')
    df = df.dropna(subset=['Date', 'Water_Indicator'])
    df['Pompa'] = df['Pompa'].astype(str).str.strip()
    return df


# ============================================================================
# ANOMALY DETECTION & PREPROCESSING ENGINE  — PATCHED v2
# ============================================================================
#
# ROOT CAUSE FIXES vs original:
#
#  BUG 1 — Stats contamination
#    BEFORE: avg_normal/std_normal/IQR dihitung dari SEMUA pos_vals termasuk
#            spike. Jika ada 1 hari spike 5000 m³, Q3 & std ikut naik sehingga
#            outlier_upper bisa menjadi ribuan → spike tidak terdeteksi.
#    FIX   : Two-pass robust stats (_robust_stats). Pass-1 buang nilai >
#            GROSS_OUTLIER_MULT × rough_median. Pass-2 hitung Q1/Q3/IQR
#            dari data bersih.
#
#  BUG 2 — outlier_upper dipilih dengan max() bukan min()
#    BEFORE: outlier_upper = max(iqr_upper, p99_upper, avg_normal * 10)
#            → selalu memilih batas PALING LONGGAR, spike lolos semua.
#    FIX   : Gunakan clean IQR fence langsung tanpa max() antar kandidat.
#
#  BUG 3 — Tidak ada "physical plausibility" per hari
#    BEFORE: Tidak ada deteksi berbasis konteks lokal (hari-hari sekitar).
#    FIX   : _rolling_clean_median() → ekspektasi lokal tiap hari (rolling
#            window 7 hari, outlier dibuang). Jika usage > PHYSICAL_MAX_MULT
#            × ekspektasi lokal → SPIKE EKSTREM, tidak peduli z-score.
#
#  BUG 4 — Digit prefix detection tidak cek digit-magnitude
#    BEFORE: Hanya coba strip prefix jika hasilnya match dalam tolerance.
#    FIX   : Tambah cek: jika indikator punya jauh lebih banyak digit dari
#            referensi, coba strip exact extra digits.
# ============================================================================

# Tuning knobs — ubah di sini jika perlu
PHYSICAL_MAX_MULT   = 10   # usage > 10× ekspektasi lokal = tidak wajar secara fisik
ROLLING_WINDOW_DAYS = 7    # hari tiap sisi untuk rolling clean median
GROSS_OUTLIER_MULT  = 30   # pass-1 cap: buang nilai > 30× rough median sebelum hitung stats
MIN_ABSOLUTE_SPIKE  = 50   # usage di bawah nilai ini TIDAK pernah dianggap spike,
                           # tidak peduli berapa× ekspektasi lokal.
                           # Mencegah unit intermittent (fin washing, dll) over-flagged
                           # karena median-nya sangat kecil (misal 0.1 m³).
                           # Aman untuk meter desimal: 50 m³/hari masih sangat besar
                           # bahkan untuk meter yang biasanya baca 0.xx m³/hari.


def _robust_stats(pos_arr: np.ndarray) -> tuple:
    """
    Hitung (clean_median, clean_std, iqr_upper) menggunakan two-pass outlier removal
    sehingga spike tidak mengkontaminasi baseline stats.

    Pass 1 — rough median, buang gross outlier (> GROSS_OUTLIER_MULT × rough_median)
    Pass 2 — hitung Q1/Q3/IQR dan std dari data bersih
    """
    arr = pos_arr[~np.isnan(pos_arr)]
    arr = arr[arr > 0]
    if len(arr) == 0:
        return 0.0, 1.0, float('inf')

    # Pass 1: buang gross outlier
    rough_med = float(np.median(arr))
    if rough_med > 0:
        clean = arr[arr <= rough_med * GROSS_OUTLIER_MULT]
    else:
        clean = arr
    if len(clean) == 0:
        clean = arr

    # Pass 2: stats dari data bersih
    clean_median = float(np.median(clean))
    clean_std    = float(np.std(clean))
    if clean_std == 0 or np.isnan(clean_std):
        clean_std = max(clean_median * 0.1, 1.0)

    if len(clean) >= 4:
        Q1  = float(np.percentile(clean, 25))
        Q3  = float(np.percentile(clean, 75))
        IQR = Q3 - Q1
        if IQR == 0:
            iqr_upper = clean_median * 5
        else:
            iqr_upper = Q3 + 3 * IQR
        # Minimal fence = 3× median (jangan over-flag unit yang stabil)
        iqr_upper = max(iqr_upper, clean_median * 3)
    else:
        iqr_upper = clean_median * 5 if clean_median > 0 else float('inf')

    return clean_median, clean_std, iqr_upper


def _rolling_clean_median(usage_arr: np.ndarray,
                           window: int = ROLLING_WINDOW_DAYS) -> np.ndarray:
    """
    Hitung rolling median harian dari usage, dengan mengabaikan gross outlier
    di dalam window. Hasilnya adalah "ekspektasi fisik lokal" tiap hari yang
    tidak terpengaruh spike di hari sekitarnya.
    """
    pos_vals = usage_arr[usage_arr > 0]
    rough_med = float(np.nanmedian(pos_vals)) if len(pos_vals) > 0 else 1.0
    spike_cap = rough_med * GROSS_OUTLIER_MULT

    result = np.full(len(usage_arr), rough_med)
    for i in range(len(usage_arr)):
        start  = max(0, i - window)
        end    = min(len(usage_arr), i + window + 1)
        w_vals = usage_arr[start:end]
        clean  = w_vals[(w_vals > 0) & (w_vals <= spike_cap)]
        if len(clean) >= 2:
            result[i] = float(np.median(clean))
        elif rough_med > 0:
            result[i] = rough_med
        # Mencegah physical_cap = 10 × 0.1 = 1 m³ pada unit intermittent
    return result


def _try_strip_prefix(indicator_val: float, reference_val: float,
                      tolerance: float = 0.10) -> tuple:
    """
    Coba strip 1–4 digit terdepan dari indicator_val.
    Returns (corrected_value, n_digits_stripped) jika cocok dalam tolerance,
    (None, 0) jika tidak.
    Contoh: 203492 → 3492, 103484 → 3484, 1003492 → 3492
    """
    if reference_val <= 0 or pd.isna(indicator_val) or pd.isna(reference_val):
        return None, 0
    s = str(int(round(abs(indicator_val))))
    for n in range(1, min(5, len(s))):
        candidate = int(s[n:])
        if candidate <= 0:
            continue
        if abs(candidate - reference_val) / max(reference_val, 1) <= tolerance:
            return float(candidate), n
    return None, 0


def detect_anomalies(df: pd.DataFrame) -> pd.DataFrame:
    """
    PATCHED v2 — Contamination-resistant anomaly detection.

    Pipeline per pompa:
      1. Two-pass robust stats (immune dari kontaminasi spike)
      2. Rolling clean median sebagai ekspektasi fisik lokal per hari
      3. PHYSICAL CAP: usage > PHYSICAL_MAX_MULT × ekspektasi lokal → SPIKE EKSTREM
         (menangkap 5031 m³/hari ketika normal 30 m³/hari, tidak peduli z-score)
      4. IQR fence dari clean stats (bukan max antar kandidat)
      5. Digit-prefix detection + digit-magnitude check
      6. Semua klasifikasi lain (negatif, meter reset, dll.) tetap sama
    """
    df = df.copy().sort_values(['Pompa', 'Date']).reset_index(drop=True)
    df['usage_raw'] = df.groupby('Pompa')['Water_Indicator'].diff()

    results = []
    for pompa, grp in df.groupby('Pompa'):
        grp   = grp.copy().sort_values('Date').reset_index(drop=True)
        ind   = grp['Water_Indicator'].values
        usage = grp['usage_raw'].copy()
        u     = usage.values

        # ── Robust baseline stats (two-pass, spike-immune) ────────────────
        all_u_arr = np.array([float(v) if not pd.isna(v) else np.nan for v in u])
        pos_arr   = all_u_arr[all_u_arr > 0]
        pos_arr   = pos_arr[~np.isnan(pos_arr)]

        clean_median, clean_std, iqr_upper = _robust_stats(pos_arr)

        # ── Per-day physical expectation (rolling clean median) ───────────
        all_u_safe     = np.where(np.isnan(all_u_arr), 0.0, all_u_arr)
        local_expected = _rolling_clean_median(all_u_safe)
        physical_cap   = local_expected * PHYSICAL_MAX_MULT   # per-day hard cap

        grp['avg_normal']    = round(clean_median, 1)
        grp['std_normal']    = round(clean_std, 1)
        grp['outlier_upper'] = round(iqr_upper, 1)
        grp['z_score']       = ((usage - clean_median) / clean_std).round(1)
        grp['faktor_kali']   = np.where(
            (usage > 0) & (clean_median > 0),
            (usage / clean_median).round(1), np.nan
        )

        anomaly_type      = ['NORMAL'] * len(grp)
        anomaly_reason    = [''] * len(grp)
        digit_prefix_flag = [False] * len(grp)
        corrected_ind     = [None] * len(grp)

        z = grp['z_score'].values

        # ── Pass 1: Digit-Prefix & Digit-Magnitude Detection ─────────────
        clean_indicators = []
        for i in range(len(grp)):
            iv = float(ind[i]) if not pd.isna(ind[i]) else None

            if len(clean_indicators) >= 2:
                ref = float(np.median(clean_indicators[-5:]))

                # A) Standard prefix strip
                corrected, n_stripped = _try_strip_prefix(iv, ref) if iv is not None else (None, 0)

                # B) Digit-magnitude check: indikator punya terlalu banyak digit vs ref
                #    Contoh: ref=3492, iv=13492 → strip 1 digit depan
                if corrected is None and iv is not None and ref > 0 and iv > 0:
                    iv_digits  = len(str(int(abs(iv))))
                    ref_digits = len(str(int(abs(ref))))
                    if iv_digits > ref_digits + 1:
                        extra_n   = iv_digits - ref_digits
                        s_iv      = str(int(abs(iv)))
                        candidate = int(s_iv[extra_n:]) if len(s_iv) > extra_n else 0
                        if candidate > 0 and abs(candidate - ref) / max(ref, 1) <= 0.15:
                            corrected  = float(candidate)
                            n_stripped = extra_n

                if corrected is not None and n_stripped >= 1:
                    digit_prefix_flag[i] = True
                    corrected_ind[i]     = corrected
                    anomaly_type[i]      = 'DIGIT PREFIX ERROR'
                    anomaly_reason[i]    = (
                        f'Kamera Salah Baca — indikator {int(iv):,} '
                        f'seharusnya {int(corrected):,} '
                        f'({n_stripped} digit ekstra di depan)'
                    )
                    clean_indicators.append(corrected)
                    continue

            if iv is not None:
                clean_indicators.append(float(iv))

        # ── Pass 2: Usage-based Anomaly Classification ────────────────────
        for i in range(len(grp)):
            if digit_prefix_flag[i]:
                continue

            v       = float(u[i]) if not pd.isna(u[i]) else 0.0
            zi      = float(z[i]) if not pd.isna(z[i]) else 0.0
            phys_i  = float(physical_cap[i])
            loc_exp = float(local_expected[i])
            faktor  = round(v / loc_exp, 1) if loc_exp > 0 and v > 0 else 0

            # ── PHYSICAL CAP (prioritas tertinggi untuk positif) ──────────
            # Tangkap kasus seperti 5031 m³/hari ketika normal ~30 m³/hari.
            # GUARD: jika usage secara absolut masih kecil (< MIN_ABSOLUTE_SPIKE),
            # TIDAK flagging — unit intermittent seperti fin washing bisa punya
            # median ~0.1 m³ sehingga 11 m³ terlihat "167× lokal" padahal wajar.
            if v > 0 and v > phys_i and loc_exp > 0 and v >= MIN_ABSOLUTE_SPIKE:
                anomaly_type[i]   = 'SPIKE EKSTREM'
                anomaly_reason[i] = (
                    f'Tidak Wajar Secara Fisik — usage {v:,.0f} m³/hari adalah '
                    f'{faktor}× ekspektasi lokal {loc_exp:.1f} m³/hari '
                    f'(batas fisik {phys_i:,.0f} m³). '
                    f'Kemungkinan: kamera salah baca digit, tidak ada gambar, atau data corrupt.'
                )
                continue

            # ── IQR Outlier (clean stats, bukan kontaminasi) ──────────────
            # GUARD sama: jangan flag nilai kecil yang secara absolut masih wajar
            if v > iqr_upper and iqr_upper < float('inf') and v >= MIN_ABSOLUTE_SPIKE:
                faktor2 = round(v / clean_median, 1) if clean_median > 0 else 0
                anomaly_type[i]   = 'SPIKE EKSTREM'
                anomaly_reason[i] = (
                    f'Outlier IQR (clean stats) — usage {v:,.0f} m³ melebihi batas wajar '
                    f'{iqr_upper:,.0f} m³ ({faktor2}× median bersih {clean_median:.1f} m³/hari). '
                    f'Kemungkinan: kamera salah baca, tidak ada gambar, atau data error.'
                )
                continue

            # ── Pergantian Flowmeter via digit-count drop ─────────────────
            if v < 0 and i > 0:
                prev_ind   = float(ind[i-1]) if not pd.isna(ind[i-1]) else 0
                curr_ind   = float(ind[i])   if not pd.isna(ind[i])   else 0
                if prev_ind > 0 and curr_ind >= 0:
                    prev_digits = len(str(int(abs(prev_ind))))
                    curr_digits = len(str(int(abs(curr_ind)))) if curr_ind > 0 else 1
                    digit_drop  = prev_digits - curr_digits
                    next_inds   = [float(ind[j]) for j in range(i+1, min(i+4, len(grp)))
                                   if not pd.isna(ind[j]) and float(ind[j]) > 0]
                    next_small  = sum(1 for ni in next_inds
                                      if len(str(int(ni))) <= curr_digits + 1)
                    if digit_drop >= 2 and next_small >= 2:
                        anomaly_type[i]   = 'PERGANTIAN FLOWMETER'
                        anomaly_reason[i] = (
                            f'Pergantian Flowmeter — indikator turun dari '
                            f'{int(prev_ind):,} ({prev_digits} digit) ke '
                            f'{int(curr_ind):,} ({curr_digits} digit), '
                            f'hari berikutnya tetap kecil (meter baru mulai dari 0)'
                        )
                        continue

            # Pergantian Flowmeter — large drop fallback
            if v < -1000:
                anomaly_type[i]   = 'PERGANTIAN FLOWMETER'
                anomaly_reason[i] = 'Pergantian Flowmeter — indikator turun drastis (reset ke angka kecil)'

            elif v < -100 and zi < -10:
                anomaly_type[i]   = 'INPUT ERROR / METER RESET'
                anomaly_reason[i] = 'Kemungkinan Kamera Terbalik atau Error Input Besar'

            elif -500 <= v < -50 and zi < -1:
                near_vals = [u[j] for j in range(max(0,i-5), min(len(u),i+5))
                             if j != i and not pd.isna(u[j])]
                similar   = sum(1 for nv in near_vals if abs(nv - v) < 20)
                anomaly_type[i]   = 'NILAI NEGATIF'
                anomaly_reason[i] = (
                    'Kamera Membaca Nama Pompa — nilai negatif berulang'
                    if similar >= 2 else
                    'Kamera Gagal / Salah Membaca Flowmeter'
                )

            elif -50 <= v < 0:
                anomaly_type[i]   = 'NILAI NEGATIF'
                anomaly_reason[i] = 'Angka Flowmeter Rolling — digit transisi antar angka'

            elif v > 0 and zi > 3 and i > 0 and (u[i-1] if not pd.isna(u[i-1]) else 0) < -50:
                if v >= MIN_ABSOLUTE_SPIKE:
                    anomaly_type[i]   = 'SPIKE'
                    anomaly_reason[i] = 'Anomali Akibat Pembacaan Salah Sebelumnya'

            # z-score spike: hanya flag jika nilai absolut juga besar
            # Unit intermittent (median ~0.1 m³) akan punya z-score tinggi
            # untuk nilai yang sebenarnya wajar (3-15 m³).
            elif zi > 20 and v >= MIN_ABSOLUTE_SPIKE:
                anomaly_type[i]   = 'SPIKE EKSTREM'
                anomaly_reason[i] = 'Tidak Ada Gambar / Kamera Gagal Total — spike sangat ekstrem'

            elif zi > 10 and v >= MIN_ABSOLUTE_SPIKE:
                anomaly_type[i]   = 'SPIKE TINGGI'
                anomaly_reason[i] = 'Kamera Gagal / Salah Membaca Flowmeter — nilai jauh di atas normal'

            elif zi > 5 and v >= MIN_ABSOLUTE_SPIKE:
                anomaly_type[i]   = 'SPIKE'
                anomaly_reason[i] = 'Flowmeter Kotor / Kamera Kurang Akurat — perlu verifikasi'

        grp['anomaly_type']       = anomaly_type
        grp['anomaly_reason']     = anomaly_reason
        grp['digit_prefix_error'] = digit_prefix_flag
        grp['corrected_indicator']= corrected_ind
        grp['is_anomaly']         = grp['anomaly_type'] != 'NORMAL'
        results.append(grp)

    return pd.concat(results, ignore_index=True)


def _estimate_usage_for_spike(pos: int, indices: list, df: 'pd.DataFrame',
                               avg_n: float, ou: float) -> float:
    """
    Estimasi usage wajar untuk hari spike dengan melihat konteks sekitarnya.

    Strategi (prioritas):
    1. Kalau hari SESUDAH bukan anomali dan usage-nya wajar → pakai usage itu
       (spike hanya 1 hari, hari berikutnya normal → mirror usage hari sesudah)
    2. Kalau ada beberapa hari normal di sekitar → rata-rata usage normal lokal
    3. Fallback: avg_normal (baseline keseluruhan)

    Ini mencegah hasil jadi 0 terus-menerus setelah spike.
    """
    # Kumpulkan usage normal dari window hari sekitar (±5 hari, bukan anomali)
    local_usages = []
    for offset in [-3, -2, -1, 1, 2, 3]:
        nbr_pos = pos + offset
        if 0 <= nbr_pos < len(indices):
            nbr_idx = indices[nbr_pos]
            nbr_anom = df.at[nbr_idx, 'is_anomaly']
            nbr_raw  = df.at[nbr_idx, 'usage_raw']
            if (not nbr_anom
                    and nbr_raw is not None
                    and not pd.isna(nbr_raw)
                    and 0 < float(nbr_raw) <= (ou if ou < float('inf') else avg_n * 5)):
                local_usages.append(float(nbr_raw))

    # Prioritas 1: hari tepat sesudah normal & wajar (cermin paling akurat)
    if pos + 1 < len(indices):
        next_idx  = indices[pos + 1]
        next_anom = df.at[next_idx, 'is_anomaly']
        next_raw  = df.at[next_idx, 'usage_raw']
        if (not next_anom
                and next_raw is not None
                and not pd.isna(next_raw)
                and 0 < float(next_raw) <= (ou if ou < float('inf') else avg_n * 5)):
            return float(next_raw)

    # Prioritas 2: median dari hari normal lokal sekitar
    if local_usages:
        import numpy as _np
        return float(_np.median(local_usages))

    # Prioritas 3: fallback ke avg_normal
    return avg_n if avg_n > 0 else 1.0


def apply_preprocessing(df_annotated: pd.DataFrame,
                        strategy: str = 'clip_to_zero') -> pd.DataFrame:
    """
    Apply preprocessing strategy. DATA NEVER DELETED.
    Original value preserved in Water_Indicator_orig.

    FINAL v5: Forward-pass chain recalculation + smart spike estimation.

    Perbaikan vs v4:
    - Spike TIDAK lagi di-set usage=cap_val yang bisa terlalu tinggi
    - Spike di-estimasi berdasarkan konteks hari sekitar (hari sesudah jika normal)
    - Ini mencegah grafik "flat/0" berkepanjangan setelah spike dikoreksi

    Prinsip per tipe anomali:
      DIGIT PREFIX ERROR    → gunakan corrected_indicator (strip digit)
      PERGANTIAN FLOWMETER  → usage = 0 (hari transisi)
      SPIKE EKSTREM/TINGGI  → usage = estimasi dari konteks lokal (bukan cap_val)
      NILAI NEGATIF kecil   → usage = 0 (rolling digit transisi, wajar)
      NILAI NEGATIF besar   → usage = 0 (kamera gagal)

    Strategies:
      flag_only      — mark only, no value changes
      clip_to_zero   — koreksi berbasis forward-pass + estimasi konteks
      interpolate    — indicator anomali diganti via time interpolation
      rolling_median — indicator anomali diganti via 7-day rolling median
    """
    df = df_annotated.copy()
    df['Water_Indicator_orig'] = df['Water_Indicator']
    df['preprocessed']         = False

    has_prefix  = 'digit_prefix_error' in df.columns
    prefix_mask = df['digit_prefix_error'].fillna(False) if has_prefix else pd.Series(False, index=df.index)

    if strategy == 'flag_only':
        for idx in df[prefix_mask].index:
            corr = df.at[idx, 'corrected_indicator']
            if corr is not None and not pd.isna(corr):
                df.at[idx, 'Water_Indicator'] = corr
                df.at[idx, 'preprocessed']    = True
        return df

    # ── Per-pompa forward-pass ────────────────────────────────────────────
    for pompa, grp in df.groupby('Pompa'):
        grp_sorted = grp.sort_values('Date').copy()
        indices    = grp_sorted.index.tolist()
        idx_anom   = set(grp_sorted[grp_sorted['is_anomaly']].index)

        avg_n = float(grp_sorted['avg_normal'].iloc[0]) if 'avg_normal' in grp_sorted.columns else 0
        ou    = float(grp_sorted['outlier_upper'].iloc[0]) if 'outlier_upper' in grp_sorted.columns else float('inf')

        # prev_corrected: indicator terkoreksi hari sebelumnya (satu-satunya state)
        prev_corrected = None

        for pos, idx in enumerate(indices):
            atype  = df.at[idx, 'anomaly_type']
            raw_v  = df.at[idx, 'usage_raw']
            orig   = float(df.at[idx, 'Water_Indicator_orig'])

            if prev_corrected is None:
                prev_corrected = orig
                continue

            prev_orig  = float(df.at[indices[pos - 1], 'Water_Indicator_orig'])
            orig_usage = orig - prev_orig

            # ── Normal day ────────────────────────────────────────────────
            if idx not in idx_anom:
                new_ind = prev_corrected + orig_usage
                if abs(new_ind - orig) > 0.01:
                    df.at[idx, 'Water_Indicator'] = round(new_ind, 2)
                    df.at[idx, 'preprocessed']    = True
                prev_corrected = new_ind
                continue

            # ── Anomali ───────────────────────────────────────────────────

            # DIGIT PREFIX ERROR
            if atype == 'DIGIT PREFIX ERROR':
                corr_ind = df.at[idx, 'corrected_indicator']
                if corr_ind is not None and not pd.isna(corr_ind):
                    usage_corr = max(0.0, float(corr_ind) - prev_orig)
                    new_ind    = prev_corrected + usage_corr
                    df.at[idx, 'Water_Indicator'] = round(new_ind, 2)
                    df.at[idx, 'preprocessed']    = True
                    prev_corrected = new_ind
                # else: prev_corrected unchanged (usage=0)
                continue

            # PERGANTIAN FLOWMETER → transition day, usage = 0
            elif atype == 'PERGANTIAN FLOWMETER':
                df.at[idx, 'Water_Indicator'] = round(prev_corrected, 2)
                df.at[idx, 'preprocessed']    = True
                # prev_corrected unchanged

            # SPIKE positif → estimasi usage dari konteks lokal
            elif (raw_v is not None and not pd.isna(raw_v) and float(raw_v) > 0
                  and atype in ('SPIKE EKSTREM', 'SPIKE TINGGI', 'SPIKE')):
                est_usage = _estimate_usage_for_spike(pos, indices, df, avg_n, ou)
                new_ind   = prev_corrected + est_usage
                df.at[idx, 'Water_Indicator'] = round(new_ind, 2)
                df.at[idx, 'preprocessed']    = True
                prev_corrected = new_ind

            # NILAI NEGATIF → usage = 0
            elif raw_v is not None and not pd.isna(raw_v) and float(raw_v) < 0:
                df.at[idx, 'Water_Indicator'] = round(prev_corrected, 2)
                df.at[idx, 'preprocessed']    = True
                # prev_corrected unchanged

            else:
                prev_corrected = orig

        # ── Strategy interpolate / rolling_median ─────────────────────────
        if strategy in ('interpolate', 'rolling_median'):
            grp_updated = df.loc[indices].sort_values('Date')
            s_orig      = grp_updated.set_index('Date')['Water_Indicator_orig'].copy().astype(float)
            s_work      = s_orig.copy()

            for idx in idx_anom:
                dv = df.at[idx, 'Date']
                s_work[dv] = np.nan

            if strategy == 'interpolate':
                try:
                    s_filled = s_work.interpolate(method='time', limit_direction='both')
                except Exception:
                    s_filled = s_work.ffill().bfill()
            else:   # rolling_median
                s_filled = s_work.copy()
                roll     = s_orig.rolling(window=7, min_periods=1, center=True).median()
                for dv in s_work[s_work.isna()].index:
                    s_filled[dv] = roll[dv]
                s_filled = s_filled.ffill().bfill()

            for idx in idx_anom:
                dv = df.at[idx, 'Date']
                if dv in s_filled.index:
                    df.at[idx, 'Water_Indicator'] = round(float(s_filled[dv]), 2)
                    df.at[idx, 'preprocessed']    = True

    return df

def get_anomaly_summary(df_annotated: pd.DataFrame) -> pd.DataFrame:
    """Return a clean summary DataFrame of all detected anomalies."""
    anom = df_annotated[df_annotated['is_anomaly']].copy()
    if anom.empty:
        return pd.DataFrame()
    anom['usage_raw'] = anom['usage_raw'].round(2)
    anom['z_score']   = anom['z_score'].round(1)
    cols = ['Date', 'Pompa', 'Location', 'Water_Indicator', 'corrected_indicator',
            'usage_raw', 'avg_normal', 'z_score',
            'anomaly_type', 'anomaly_reason', 'digit_prefix_error', 'preprocessed']
    cols = [c for c in cols if c in anom.columns]
    return anom[cols].sort_values(['anomaly_type', 'z_score']).reset_index(drop=True)


# ============================================================================
# DATA PROCESSING (uses preprocessing pipeline)
# ============================================================================

def process_data(df: pd.DataFrame, selected_prods: list,
                 dedup_method: str,
                 preprocess_strategy: str = 'flag_only'):
    # max_spike: hardcoded sangat tinggi — anomaly detection sudah handle
    # spike via MIN_ABSOLUTE_SPIKE, tidak perlu user-facing clip lagi
    max_spike = 999_999.0
    if df.empty or not selected_prods:
        return None

    filt = df[df['Pompa'].isin(selected_prods)].copy()
    if filt.empty:
        return None

    # ── Step 1: Dedup ──────────────────────────────────────────────────────
    if dedup_method == 'First':
        dedup = filt.groupby(['Date', 'Pompa'])['Water_Indicator'].first().reset_index()
    elif dedup_method == 'Last':
        dedup = filt.groupby(['Date', 'Pompa'])['Water_Indicator'].last().reset_index()
    elif dedup_method == 'Max':
        dedup = filt.groupby(['Date', 'Pompa'])['Water_Indicator'].max().reset_index()
    else:
        dedup = filt.groupby(['Date', 'Pompa'])['Water_Indicator'].mean().reset_index()

    loc_map = filt.groupby('Pompa')['Location'].first().to_dict()
    dedup['Location'] = dedup['Pompa'].map(loc_map)

    # ── Step 2: Anomaly Detection ──────────────────────────────────────────
    if preprocess_strategy != 'none':
        df_ann   = detect_anomalies(dedup)
        df_clean = apply_preprocessing(df_ann, strategy=preprocess_strategy)
        dedup_use = df_clean[['Date', 'Pompa', 'Location', 'Water_Indicator']].copy()
    else:
        df_ann    = detect_anomalies(dedup)
        df_clean  = df_ann.copy()
        df_clean['preprocessed'] = False
        dedup_use = dedup

    # ── Step 3: Pivot & consumption diff from CLEAN indicators ────────────
    pivot = (dedup_use.pivot(index='Date', columns='Pompa', values='Water_Indicator')
                       .sort_index())

    cons_raw = pivot.diff()
    cons_raw.iloc[0] = 0

    # ── Step 4: Gap-aware daily averaging ────────────────────────────────
    # Kalau ada tanggal yang skip (misal data bulan depan baru masuk setelah
    # 29 hari kosong), pivot.diff() akan mengembalikan TOTAL usage N hari
    # sekaligus di 1 baris — terlihat seperti spike padahal normal.
    #
    # Fix: hitung gap (selisih hari antar tanggal berurutan). Kalau gap > 1,
    # bagi usage secara rata ke setiap hari dalam gap tersebut.
    #
    # Contoh: 2025-03-13 → 2025-04-11 (gap 29 hari), usage 749 m³
    #   → tiap hari dalam range itu: 749 / 29 ≈ 25.8 m³/hari  ✓
    #   vs tanpa gap handling: 749 m³ di 1 hari → kena flag spike  ✗

    GAP_THRESHOLD_DAYS = 2   # gap ≥ 2 hari → distribusi rata

    # Hitung gap antar tanggal yang benar-benar ada di data (bukan kalender penuh)
    dates_present = pivot.index  # DatetimeIndex, sudah sorted
    date_gaps     = pd.Series(dates_present).diff().dt.days.fillna(1).values  # gap[i] = tanggal[i] - tanggal[i-1]

    cons = cons_raw.copy()
    for i, gap in enumerate(date_gaps):
        if gap >= GAP_THRESHOLD_DAYS:
            # Bagi usage secara rata ke tiap hari dalam gap
            cons.iloc[i] = cons_raw.iloc[i] / gap

    # Final safety net + fillna:
    # - clip lower=0: usage tidak boleh negatif setelah preprocessing
    # - fillna(0): unit yang tidak punya data hari tertentu → 0 (tidak putus di chart)
    # - max_spike sudah tidak relevan (anomaly detection yang handle)
    cons = cons.clip(lower=0).fillna(0)

    return pivot, cons, dedup_use, loc_map, df_ann, df_clean


# ============================================================================
# HELPERS
# ============================================================================

def fmt_date(d) -> str:
    try:
        return pd.Timestamp(d).strftime('%d %b %Y')
    except Exception:
        return str(d)


def kpi_card_html(prod: str, area: str, total: float, avg: float,
                  max_v: float, pct: float, color: str) -> str:
    r, g, b = int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)
    return f"""
    <div style="background:white;border-radius:14px;overflow:hidden;
                box-shadow:0 1px 6px rgba(0,0,0,0.08);border-top:4px solid {color};
                padding:16px 18px;font-family:Inter,sans-serif;margin-bottom:4px;">
      <div style="font-size:11px;font-weight:700;color:{color};letter-spacing:1px;
                  text-transform:uppercase;margin-bottom:4px;">{prod}</div>
      <div style="font-size:11px;color:#64748b;margin-bottom:10px;
                  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">{area}</div>
      <div style="font-size:34px;font-weight:700;color:#0f172a;line-height:1;">{total:,.1f}</div>
      <div style="font-size:12px;color:#94a3b8;margin-bottom:10px;">m³ total</div>
      <div style="display:flex;gap:10px;font-size:11px;margin-bottom:10px;">
        <div style="flex:1;background:#f8fafc;border-radius:8px;padding:6px 8px;">
          <div style="color:#64748b;">Avg / day</div>
          <div style="font-weight:700;color:#0f172a;">{avg:.1f} m³</div>
        </div>
        <div style="flex:1;background:#f8fafc;border-radius:8px;padding:6px 8px;">
          <div style="color:#64748b;">Max / day</div>
          <div style="font-weight:700;color:#0f172a;">{max_v:.1f} m³</div>
        </div>
      </div>
      <div style="height:6px;background:#f0f0f0;border-radius:99px;overflow:hidden;">
        <div style="height:100%;width:{min(pct,100):.1f}%;
                    background:linear-gradient(90deg,rgba({r},{g},{b},0.4),{color});
                    border-radius:99px;"></div>
      </div>
      <div style="font-size:10px;color:#94a3b8;margin-top:4px;">{pct:.1f}% of total</div>
    </div>"""


# ============================================================================
# CHART FUNCTIONS
# ============================================================================

def make_pie(cons_totals: pd.Series, loc_map: dict) -> go.Figure:
    colors = [get_unit_color(p) for p in cons_totals.index]

    def short_area(prod):
        a = loc_map.get(prod, '')
        return a.replace(f'{prod} - ', '').replace(prod, '').strip(' -')

    prods   = cons_totals.index.tolist()
    vals    = cons_totals.values
    areas   = [short_area(p) for p in prods]
    total   = vals.sum() if vals.sum() > 0 else 1
    pcts    = [v / total * 100 for v in vals]

    hover_texts  = [
        f"<b>{p}</b><br>{a}<br>{v:,.1f} m³  ({pct:.1f}%)"
        for p, a, v, pct in zip(prods, areas, vals, pcts)
    ]
    slice_labels = [
        f"<b>{p}</b><br>{a}<br>{v:,.1f} m³<br>{pct:.1f}%"
        for p, a, v, pct in zip(prods, areas, vals, pcts)
    ]

    fig = go.Figure(go.Pie(
        labels=prods,
        values=vals,
        text=slice_labels,
        hovertext=hover_texts,
        hoverinfo='text',
        texttemplate='%{text}',
        textfont=dict(size=10),
        marker=dict(colors=colors, line=dict(color='white', width=2)),
        hole=0.35,
        insidetextorientation='radial',
    ))
    fig.update_layout(
        title=dict(text='Consumption Distribution by Unit', font=dict(size=15, color='#1e293b')),
        paper_bgcolor='white',
        showlegend=True,
        legend=dict(
            orientation='h',
            yanchor='bottom', y=-0.25,
            xanchor='center', x=0.5,
            font=dict(size=10),
            itemsizing='constant',
        ),
        height=480, margin=dict(l=120, r=120, t=50, b=80)
    )
    return fig


def make_bar_total(cons_totals: pd.Series, loc_map: dict) -> go.Figure:
    """For Streamlit display — outside labels, wide margin."""
    prods  = cons_totals.index.tolist()
    vals   = cons_totals.values
    colors = [get_unit_color(p) for p in prods]

    def short_area(prod):
        a = loc_map.get(prod, '')
        return a.replace(f'{prod} - ', '').replace(prod, '').strip(' -')

    areas = [short_area(p) for p in prods]
    y_labels    = [f"{p} — {a}" for p, a in zip(prods, areas)]
    hover_texts = [f"<b>{p}</b><br>{a}<br>{v:,.1f} m³" for p, a, v in zip(prods, areas, vals)]

    fig = go.Figure(go.Bar(
        x=vals,
        y=y_labels,
        orientation='h',
        marker=dict(color=colors, line=dict(color='white', width=1)),
        text=[f"{v:,.0f} m³" for v in vals],
        textposition='outside',
        hovertext=hover_texts,
        hoverinfo='text',
    ))
    x_max = max(vals) * 1.30 if len(vals) else 1
    fig.update_layout(
        title=dict(text='Total Water Consumption by Unit (m³)', font=dict(size=15, color='#1e293b')),
        paper_bgcolor='white', plot_bgcolor='white',
        height=max(380, len(prods) * 55 + 80),
        margin=dict(l=10, r=20, t=50, b=20),
        xaxis=dict(showgrid=True, gridcolor='#f1f5f9', linecolor='#e2e8f0',
                   tickfont=dict(size=10), range=[0, x_max]),
        yaxis=dict(showgrid=False, linecolor='#e2e8f0',
                   tickfont=dict(size=11), autorange='reversed'),
    )
    return fig


def make_bar_total_html(cons_totals: pd.Series, loc_map: dict) -> go.Figure:
    """For HTML export — robust horizontal bar chart, labels inside bars."""
    prods  = cons_totals.index.tolist()
    # Convert to plain Python floats to avoid numpy serialization issues
    vals   = [float(v) for v in cons_totals.values]
    colors = [get_unit_color(p) for p in prods]

    def short_area(prod):
        a = loc_map.get(prod, '')
        return a.replace(f'{prod} - ', '').replace(prod, '').strip(' -')

    areas       = [short_area(p) for p in prods]
    hover_texts = [f"<b>{p}</b><br>{a}<br>{v:,.1f} m³" for p, a, v in zip(prods, areas, vals)]

    traces = []
    for i, (prod, val, color, area, hover) in enumerate(zip(prods, vals, colors, areas, hover_texts)):
        traces.append(go.Bar(
            x=[val],
            y=[prod],
            orientation='h',
            name=prod,
            showlegend=False,
            marker=dict(color=color, line=dict(color='white', width=1)),
            text=[f"{area} — {val:,.0f} m³"],
            textposition='inside',
            insidetextanchor='start',
            textfont=dict(size=11, color='white'),
            hovertext=[hover],
            hoverinfo='text',
        ))

    fig = go.Figure(data=traces)
    x_max = max(vals) * 1.08 if vals else 1

    fig.update_layout(
        title=dict(text='Total Water Consumption by Unit (m³)', font=dict(size=15, color='#1e293b')),
        paper_bgcolor='white', plot_bgcolor='white',
        barmode='overlay',
        height=max(380, len(prods) * 58 + 80),
        margin=dict(l=90, r=30, t=50, b=20),
        xaxis=dict(
            showgrid=True, gridcolor='#f1f5f9', linecolor='#e2e8f0',
            tickfont=dict(size=10),
            range=[0, x_max],
            autorange=False,
        ),
        yaxis=dict(
            showgrid=False, linecolor='#e2e8f0',
            tickfont=dict(size=12, color='#1e293b'),
            autorange='reversed',
            automargin=True,
        ),
        uniformtext=dict(minsize=8, mode='hide'),
    )
    return fig


def make_line_daily(cons: pd.DataFrame, loc_map: dict, selected_prods: list) -> go.Figure:
    """Dual y-axis linear scale: large units on left axis, small units on right axis."""
    from plotly.subplots import make_subplots

    cols = [p for p in selected_prods if p in cons.columns]
    if not cols:
        return go.Figure()

    x_lbl = [fmt_date(d) for d in cons.index]

    totals  = {p: float(cons[p].sum()) for p in cols}
    max_tot = max(totals.values()) if totals else 1
    large   = [p for p in cols if totals[p] >= max_tot * 0.10]
    small   = [p for p in cols if totals[p] <  max_tot * 0.10]

    fig = make_subplots(specs=[[{"secondary_y": bool(small)}]])

    for prod in large:
        color = get_unit_color(prod)
        area  = loc_map.get(prod, '')
        fig.add_trace(go.Scatter(
            x=x_lbl,
            y=[float(v) for v in cons[prod].tolist()],
            mode='lines+markers', name=prod,
            line=dict(color=color, width=2),
            marker=dict(size=5, color=color),
            hovertemplate='<b>' + prod + '</b> — ' + area + '<br>%{x}<br>%{y:,.2f} m³<extra></extra>'
        ), secondary_y=False)

    for prod in small:
        color = get_unit_color(prod)
        area  = loc_map.get(prod, '')
        fig.add_trace(go.Scatter(
            x=x_lbl,
            y=[float(v) for v in cons[prod].tolist()],
            mode='lines+markers', name=prod + ' ▷',
            line=dict(color=color, width=2, dash='dot'),
            marker=dict(size=5, color=color, symbol='diamond'),
            hovertemplate='<b>' + prod + '</b> — ' + area + '<br>%{x}<br>%{y:,.2f} m³ (right)<extra></extra>'
        ), secondary_y=True)

    title_txt = 'Daily Consumption Trend by Unit (m³/day)' + (' — Dual Axis' if small else '')

    fig.update_layout(
        title=dict(text=title_txt, font=dict(size=15, color='#1e293b')),
        paper_bgcolor='white', plot_bgcolor='white',
        height=450, margin=dict(l=70, r=70, t=60, b=60),
        xaxis=dict(showgrid=False, linecolor='#e2e8f0',
                   tickfont=dict(size=10), tickangle=-30),
        legend=dict(orientation='h', yanchor='bottom', y=1.02,
                    xanchor='right', x=1, font=dict(size=10)),
        hovermode='x unified'
    )
    fig.update_yaxes(
        title_text='m³/day',
        showgrid=True, gridcolor='#f1f5f9',
        rangemode='tozero',
        secondary_y=False
    )
    if small:
        fig.update_yaxes(
            title_text='m³/day (small units ▷)',
            showgrid=False,
            rangemode='tozero',
            secondary_y=True
        )
        fig.add_annotation(
            text='Dashed lines (◆) use right y-axis',
            xref='paper', yref='paper', x=0.01, y=1.08,
            showarrow=False, font=dict(size=10, color='#94a3b8'),
            align='left'
        )
    return fig



def make_stacked_bar(cons: pd.DataFrame, loc_map: dict, selected_prods: list) -> go.Figure:
    fig = go.Figure()
    x_lbl = [fmt_date(d) for d in cons.index]
    for prod in selected_prods:
        if prod not in cons.columns:
            continue
        color = get_unit_color(prod)
        fig.add_trace(go.Bar(
            x=x_lbl, y=cons[prod], name=prod,
            marker_color=color,
            hovertemplate='<b>' + prod + '</b><br>%{x}<br>%{y:,.2f} m³<extra></extra>'
        ))
    fig.update_layout(
        title=dict(text='Daily Consumption — Stacked (m³)', font=dict(size=15, color='#1e293b')),
        barmode='stack', paper_bgcolor='white', plot_bgcolor='white',
        height=420, margin=dict(l=60, r=20, t=50, b=60),
        xaxis=dict(showgrid=False, linecolor='#e2e8f0', tickfont=dict(size=10), tickangle=-30),
        yaxis=dict(showgrid=True, gridcolor='#f1f5f9',
                   title=dict(text='m³/day', font=dict(size=12))),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
        hovermode='x unified'
    )
    return fig


def make_heatmap(cons: pd.DataFrame, selected_prods: list) -> go.Figure:
    """FIX: Heatmap was blank because zmin=zmax when one value dominates.
    Use per-row normalization so all units are visible."""
    cols = [p for p in selected_prods if p in cons.columns]
    if not cols:
        return go.Figure()

    z_raw = cons[cols].T.values.astype(float)
    x     = [fmt_date(d) for d in cons.index]

    # Normalize each row (unit) independently so small units are visible
    z_norm = np.zeros_like(z_raw)
    for i in range(len(z_raw)):
        row_max = z_raw[i].max()
        if row_max > 0:
            z_norm[i] = z_raw[i] / row_max * 100
        else:
            z_norm[i] = z_raw[i]

    # Convert to Python lists for reliable JSON serialization in HTML export
    z_norm_list = z_norm.tolist()
    z_raw_list  = z_raw.tolist()

    # Build hover text with actual values
    hover = []
    for i, prod in enumerate(cols):
        row_hover = []
        for j, date in enumerate(x):
            actual = z_raw[i][j]
            row_hover.append(f"<b>{prod}</b><br>{date}<br>{actual:,.2f} m³")
        hover.append(row_hover)

    fig = go.Figure(go.Heatmap(
        z=z_norm_list,
        x=x,
        y=cols,
        customdata=z_raw_list,
        colorscale='Blues',
        zmin=0,
        zmax=100,
        hovertext=hover,
        hoverinfo='text',
        colorbar=dict(
            title=dict(text='Relative Intensity (%)', side='right'),
            ticksuffix='%'
        )
    ))
    fig.update_layout(
        title=dict(text='Daily Consumption Heatmap (relative intensity per unit)', font=dict(size=15, color='#1e293b')),
        paper_bgcolor='white', plot_bgcolor='white',
        height=max(300, len(cols) * 45 + 100),
        margin=dict(l=80, r=80, t=60, b=60),
        xaxis=dict(tickfont=dict(size=9), tickangle=-30),
        yaxis=dict(tickfont=dict(size=11))
    )
    return fig


def make_cumulative(cons: pd.DataFrame, loc_map: dict, selected_prods: list) -> go.Figure:
    """Use log scale so all units are visible. Each trace fills to zero independently."""
    fig = go.Figure()
    x_lbl = [fmt_date(d) for d in cons.index]
    for prod in selected_prods:
        if prod not in cons.columns:
            continue
        color = get_unit_color(prod)
        cumul = cons[prod].cumsum().replace(0, np.nan)
        r, g, b = int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)
        fig.add_trace(go.Scatter(
            x=x_lbl, y=cumul, mode='lines', name=prod,
            line=dict(color=color, width=2),
            fill='tozeroy',
            fillcolor=f'rgba({r},{g},{b},0.08)',
            hovertemplate='<b>' + prod + '</b> cumulative<br>%{x}<br>%{y:,.1f} m³<extra></extra>'
        ))
    fig.update_layout(
        title=dict(text='Cumulative Consumption (m³) — Log Scale', font=dict(size=15, color='#1e293b')),
        paper_bgcolor='white', plot_bgcolor='white',
        height=420, margin=dict(l=70, r=20, t=50, b=60),
        xaxis=dict(showgrid=False, linecolor='#e2e8f0', tickangle=-30,
                   tickfont=dict(size=10)),
        yaxis=dict(showgrid=True, gridcolor='#f1f5f9',
                   type='log',
                   title=dict(text='Cumulative m³ (log scale)', font=dict(size=12))),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
        hovermode='x unified'
    )
    return fig


def make_avg_max_bar(cons: pd.DataFrame, loc_map: dict, selected_prods: list) -> go.Figure:
    def short_area(unit):
        a = loc_map.get(unit, '')
        return a.replace(f'{unit} - ', '').replace(unit, '').strip(' -')

    units  = [u for u in selected_prods if u in cons.columns]
    labels = [f"{u} — {short_area(u)}" for u in units]
    avgs   = [float(cons[u][cons[u] > 0].mean()) if (cons[u] > 0).any() else 0 for u in units]
    maxes  = [float(cons[u][cons[u] > 0].max())  if (cons[u] > 0).any() else 0 for u in units]

    fig = go.Figure()
    fig.add_trace(go.Bar(
        name='Avg / day',
        x=labels, y=avgs,
        marker_color='#0ea5e9',
        text=[f"{v:.1f}" for v in avgs], textposition='outside',
        hovertemplate='<b>%{x}</b><br>Avg/day: %{y:,.2f} m³<extra></extra>'
    ))
    fig.add_trace(go.Bar(
        name='Max / day',
        x=labels, y=maxes,
        marker_color='#ef4444',
        text=[f"{v:.1f}" for v in maxes], textposition='outside',
        hovertemplate='<b>%{x}</b><br>Max/day: %{y:,.2f} m³<extra></extra>'
    ))
    fig.update_layout(
        title=dict(text='Avg & Max Daily Consumption by Unit (m³/day)', font=dict(size=15, color='#1e293b')),
        barmode='group', paper_bgcolor='white', plot_bgcolor='white',
        height=420, margin=dict(l=60, r=10, t=50, b=100),
        xaxis=dict(showgrid=False, linecolor='#e2e8f0', tickfont=dict(size=10), tickangle=-20),
        yaxis=dict(showgrid=True, gridcolor='#f1f5f9',
                   title=dict(text='m³/day', font=dict(size=12))),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
    )
    return fig


# ============================================================================
# PEAK DAY HELPER  ← NEW
# ============================================================================

def get_peak_day_info(cons_active: pd.DataFrame) -> tuple:
    """Return (peak_value, peak_date_str) for the day with highest total consumption."""
    if cons_active.empty:
        return 0.0, "—"
    daily_totals = cons_active.sum(axis=1)
    peak_val  = float(daily_totals.max())
    peak_date = daily_totals.idxmax()
    peak_str  = pd.Timestamp(peak_date).strftime('%d %b %Y (%A)')
    return peak_val, peak_str


# ============================================================================
# EXCEL EXPORT
# ============================================================================

def generate_excel_report(
    raw_filtered: pd.DataFrame,
    cons: pd.DataFrame,
    cons_active: pd.DataFrame,
    cons_totals: pd.Series,
    loc_map: dict,
    selected_prods: list,
    period_str: str,
    total_all: float,
    n_days: int,
) -> bytes:
    import io as _io
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    HEADER_FILL = PatternFill('solid', start_color='0C4A6E')
    HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    SUBHDR_FILL = PatternFill('solid', start_color='0EA5E9')
    SUBHDR_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    ALT_FILL    = PatternFill('solid', start_color='F0F9FF')
    TOTAL_FILL  = PatternFill('solid', start_color='0C4A6E')
    TOTAL_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    NORMAL_FONT = Font(name='Arial', size=10)
    BOLD_FONT   = Font(bold=True, name='Arial', size=10)
    thin        = Side(style='thin', color='E2E8F0')
    BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER      = Alignment(horizontal='center', vertical='center')
    LEFT        = Alignment(horizontal='left', vertical='center')

    def set_row(ws, row, values, fill=None, font=None, align=None):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            if fill:  c.fill      = fill
            if font:  c.font      = font
            if align: c.alignment = align
            c.border = BORDER

    def set_col_widths(ws, widths):
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    def short_area(prod):
        a = loc_map.get(prod, '')
        return a.replace(f'{prod} - ', '').replace(prod, '').strip(' -')

    avg_daily = float(cons_active.sum(axis=1).mean()) if len(cons_active) > 0 else 0
    peak_val, peak_date_str = get_peak_day_info(cons_active)
    active_cols = [p for p in selected_prods if p in cons.columns]

    # =========================================================================
    # SHEET 1 — Summary
    # =========================================================================
    ws1 = wb.active
    ws1.title = 'Summary'
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells('A1:F1')
    ws1['A1'].value = 'Water Consumption Report'
    ws1['A1'].font  = Font(bold=True, name='Arial', size=16, color='0C4A6E')
    ws1['A1'].alignment = CENTER
    ws1['A1'].fill  = PatternFill('solid', start_color='F0F9FF')
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells('A2:F2')
    ws1['A2'].value = f'Period: {period_str}  |  Units: {", ".join(selected_prods)}'
    ws1['A2'].font  = Font(name='Arial', size=10, color='64748B')
    ws1['A2'].alignment = CENTER
    ws1.row_dimensions[2].height = 20

    ws1.merge_cells('A4:F4')
    ws1['A4'].value = 'KEY PERFORMANCE INDICATORS'
    ws1['A4'].font  = Font(bold=True, name='Arial', size=11, color='FFFFFF')
    ws1['A4'].fill  = HEADER_FILL
    ws1['A4'].alignment = CENTER
    ws1.row_dimensions[4].height = 22

    set_row(ws1, 5, ['Metric', 'Value', 'Unit', '', 'Metric', 'Value'],
            fill=SUBHDR_FILL, font=SUBHDR_FONT, align=CENTER)

    kpis_l = [('Total Consumption', f'{total_all:,.1f}', 'm³'),
              ('Days of Data',      str(n_days),         'days'),
              ('Average / Day',     f'{avg_daily:,.1f}', 'm³/day')]
    kpis_r = [('Peak Day Value',    f'{peak_val:,.1f}',  'm³'),
              ('Peak Day Date',     peak_date_str,        ''),
              ('Units Monitored',   str(len(selected_prods)), 'units')]

    for i, ((lk, lv, lu), (rk, rv, _)) in enumerate(zip(kpis_l, kpis_r), 6):
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for col, val in enumerate([lk, lv, lu, '', rk, rv], 1):
            c = ws1.cell(row=i, column=col, value=val)
            c.font      = BOLD_FONT if col in (1, 5) else NORMAL_FONT
            c.fill      = fill
            c.border    = BORDER
            c.alignment = LEFT if col in (1, 5) else CENTER

    ws1.merge_cells('A10:G10')
    ws1['A10'].value = 'UNIT STATISTICS'
    ws1['A10'].font  = Font(bold=True, name='Arial', size=11, color='FFFFFF')
    ws1['A10'].fill  = HEADER_FILL
    ws1['A10'].alignment = CENTER
    ws1.row_dimensions[10].height = 22

    set_row(ws1, 11,
            ['Unit', 'Area', 'Total (m³)', 'Avg/day (m³)', 'Max/day (m³)', 'Min/day (m³)', '% of Total'],
            fill=SUBHDR_FILL, font=SUBHDR_FONT, align=CENTER)

    for i, prod in enumerate(selected_prods, 12):
        if prod not in cons_active.columns: continue
        d     = cons_active[prod]
        d_pos = d[d > 0]
        pct   = round(float(d.sum()) / total_all * 100, 1) if total_all > 0 else 0
        row_v = [prod, short_area(prod),
                 round(float(d.sum()), 1),
                 round(float(d_pos.mean()), 2) if len(d_pos) else 0,
                 round(float(d_pos.max()),  1) if len(d_pos) else 0,
                 round(float(d_pos.min()),  1) if len(d_pos) else 0,
                 f'{pct:.1f}%']
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for col, val in enumerate(row_v, 1):
            c = ws1.cell(row=i, column=col, value=val)
            c.font = NORMAL_FONT; c.fill = fill; c.border = BORDER
            c.alignment = LEFT if col <= 2 else CENTER

    tr = 12 + len(selected_prods)
    set_row(ws1, tr, ['TOTAL', '', f'{total_all:,.1f}', '', '', '', '100.0%'],
            fill=TOTAL_FILL, font=TOTAL_FONT, align=CENTER)
    set_col_widths(ws1, [14, 30, 14, 15, 15, 15, 12])

    # =========================================================================
    # SHEET 2 — Raw Data
    # =========================================================================
    ws2 = wb.create_sheet('Raw Data')
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells('A1:D1')
    ws2['A1'].value = f'Raw Data — {period_str}'
    ws2['A1'].font  = Font(bold=True, name='Arial', size=13, color='0C4A6E')
    ws2['A1'].alignment = CENTER
    ws2['A1'].fill  = PatternFill('solid', start_color='F0F9FF')
    ws2.row_dimensions[1].height = 28

    set_row(ws2, 2, ['Date', 'Location', 'Unit (Pompa)', 'Water Indicator (m³)'],
            fill=HEADER_FILL, font=HEADER_FONT, align=CENTER)

    raw_exp = raw_filtered.copy()
    raw_exp['Date'] = pd.to_datetime(raw_exp['Date']).dt.strftime('%d-%b-%Y')
    raw_exp = raw_exp.sort_values(['Pompa', 'Date'])

    for i, (_, row) in enumerate(raw_exp.iterrows(), 3):
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        vals = [row.get('Date',''), row.get('Location',''),
                row.get('Pompa',''), round(float(row.get('Water_Indicator', 0)), 2)]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=i, column=col, value=val)
            c.font = NORMAL_FONT; c.fill = fill; c.border = BORDER
            c.alignment = LEFT if col <= 3 else CENTER
            if col == 4: c.number_format = '#,##0.00'

    set_col_widths(ws2, [14, 32, 14, 22])

    # =========================================================================
    # SHEET 3 — Water Usage (m³)
    # =========================================================================
    ws3 = wb.create_sheet('Water Usage (m3)')
    ws3.sheet_view.showGridLines = False

    ncols = len(active_cols)
    ws3.merge_cells(f'A1:{get_column_letter(ncols + 2)}1')
    ws3['A1'].value = f'Daily Water Consumption (m³) — {period_str}'
    ws3['A1'].font  = Font(bold=True, name='Arial', size=13, color='0C4A6E')
    ws3['A1'].alignment = CENTER
    ws3['A1'].fill  = PatternFill('solid', start_color='F0F9FF')
    ws3.row_dimensions[1].height = 28

    # Area sub-header row
    for col, val in enumerate(['Date'] + [short_area(p) for p in active_cols] + ['TOTAL'], 1):
        c = ws3.cell(row=2, column=col, value=val)
        c.font = Font(italic=True, name='Arial', size=9, color='0C4A6E')
        c.fill = PatternFill('solid', start_color='E0F2FE')
        c.alignment = CENTER; c.border = BORDER

    set_row(ws3, 3, ['Date'] + active_cols + ['TOTAL'],
            fill=HEADER_FILL, font=HEADER_FONT, align=CENTER)

    cons_disp = cons.iloc[1:]
    for i, (date, row) in enumerate(cons_disp.iterrows(), 4):
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        row_total = sum(float(row.get(p, 0)) for p in active_cols)
        vals = [fmt_date(date)] + [round(float(row.get(p, 0)), 2) for p in active_cols] + [round(row_total, 2)]
        for col, val in enumerate(vals, 1):
            c = ws3.cell(row=i, column=col, value=val)
            c.font = NORMAL_FONT; c.fill = fill; c.border = BORDER
            c.alignment = LEFT if col == 1 else CENTER
            if col > 1: c.number_format = '#,##0.00'

    tr3 = 4 + len(cons_disp)
    grand = ['GRAND TOTAL'] + [round(float(cons_totals.get(p, 0)), 1) for p in active_cols] + [round(total_all, 1)]
    set_row(ws3, tr3, grand, fill=TOTAL_FILL, font=TOTAL_FONT, align=CENTER)

    # Monthly summary
    br = tr3 + 2
    ws3.merge_cells(f'A{br}:{get_column_letter(ncols+2)}{br}')
    ws3.cell(row=br, column=1).value = 'MONTHLY SUMMARY'
    ws3.cell(row=br, column=1).font  = Font(bold=True, name='Arial', size=11, color='FFFFFF')
    ws3.cell(row=br, column=1).fill  = HEADER_FILL
    ws3.cell(row=br, column=1).alignment = CENTER

    set_row(ws3, br+1, ['Month'] + active_cols + ['TOTAL'],
            fill=SUBHDR_FILL, font=SUBHDR_FONT, align=CENTER)

    tmp_c = cons.copy()
    tmp_c.index = pd.to_datetime(tmp_c.index)
    monthly = tmp_c.resample('ME').sum()
    for i, (mdate, mrow) in enumerate(monthly.iterrows(), br + 2):
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        mt = sum(float(mrow.get(p, 0)) for p in active_cols)
        mv = [mdate.strftime('%B %Y')] + [round(float(mrow.get(p,0)),1) for p in active_cols] + [round(mt,1)]
        for col, val in enumerate(mv, 1):
            c = ws3.cell(row=i, column=col, value=val)
            c.font = NORMAL_FONT; c.fill = fill; c.border = BORDER
            c.alignment = LEFT if col == 1 else CENTER
            if col > 1: c.number_format = '#,##0.0'

    set_col_widths(ws3, [14] + [13]*ncols + [13])

    # =========================================================================
    # SHEET 4 — Native Excel Charts (read from Water Usage (m3) sheet data)
    # =========================================================================
    from openpyxl.chart import (BarChart, LineChart, PieChart,
                                 AreaChart, Reference, Series)
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.label import DataLabelList

    ws4 = wb.create_sheet('Charts')
    ws4.sheet_view.showGridLines = False
    from openpyxl.chart import (BarChart, LineChart, PieChart, AreaChart, Reference)
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.label import DataLabelList
    from openpyxl.worksheet.datavalidation import DataValidation

    # ── Column widths ─────────────────────────────────────────────────────────
    # A-H (1-8) : visible summary table
    # I (9)     : spacer
    # J+ (10+)  : hidden helper data for chart references
    col_widths = {'A': 11, 'B': 26, 'C': 13, 'D': 14, 'E': 14, 'F': 11, 'G': 18, 'H': 11}
    for ltr, w in col_widths.items():
        ws4.column_dimensions[ltr].width = w
    ws4.column_dimensions['I'].width = 2
    for col in range(10, 80):
        ws4.column_dimensions[get_column_letter(col)].width = 0.1

    CHART_W        = 28
    CHART_H        = 14
    ROW_H_PT       = 15.0
    CM_PER_PT      = 0.0353
    rows_per_chart = int(CHART_H / (ROW_H_PT * CM_PER_PT)) + 5

    data_start_row = 4
    data_end_row   = tr3 - 1

    # ══════════════════════════════════════════════════════════════════════════
    # ROW 1 — Title banner
    # ══════════════════════════════════════════════════════════════════════════
    ws4.merge_cells('A1:H1')
    c = ws4.cell(row=1, column=1, value=f'Visualisations \u2014 {period_str}')
    c.font = Font(bold=True, name='Arial', size=14, color='0C4A6E')
    c.fill = PatternFill('solid', start_color='F0F9FF')
    c.alignment = CENTER
    ws4.row_dimensions[1].height = 30

    # ══════════════════════════════════════════════════════════════════════════
    # ROW 2 — "Show Location Name?" toggle  (dropdown YES / NO in cell H2)
    # ══════════════════════════════════════════════════════════════════════════
    ws4.row_dimensions[2].height = 22
    ws4.merge_cells('A2:F2')
    lbl = ws4.cell(row=2, column=1,
                   value='\u2699  Show Location Name in Chart Labels?  \u2192  change cell H2')
    lbl.font = Font(bold=True, name='Arial', size=10, color='0C4A6E')
    lbl.fill = PatternFill('solid', start_color='E0F2FE')
    lbl.alignment = LEFT

    note = ws4.cell(row=2, column=7,
                    value='YES = with area  |  NO = unit only')
    note.font = Font(italic=True, name='Arial', size=9, color='64748B')
    note.alignment = LEFT

    ws4['H2'] = 'YES'
    tc = ws4['H2']
    tc.font      = Font(bold=True, name='Arial', size=11, color='FFFFFF')
    tc.fill      = PatternFill('solid', start_color='0EA5E9')
    tc.alignment = CENTER
    tc.border    = BORDER

    dv = DataValidation(type='list', formula1='"YES,NO"', allow_blank=False)
    dv.sqref = 'H2'
    dv.showDropDown = False
    ws4.add_data_validation(dv)

    # ══════════════════════════════════════════════════════════════════════════
    # ROWS 3+ — Visible Summary Table  (cols A-H)
    # ══════════════════════════════════════════════════════════════════════════
    tbl_hdr_row = 3
    ws4.row_dimensions[tbl_hdr_row].height = 20
    ws4.merge_cells(f'A{tbl_hdr_row}:H{tbl_hdr_row}')
    th = ws4.cell(row=tbl_hdr_row, column=1,
                  value='\U0001f4ca  Data Summary  \u2014  source for charts below')
    th.font = Font(bold=True, name='Arial', size=11, color='FFFFFF')
    th.fill = HEADER_FILL; th.alignment = CENTER

    hdrs = ['Unit', 'Area / Location', 'Total (m\u00b3)', 'Avg/day (m\u00b3)',
            'Max/day (m\u00b3)', '% of Total', 'Active Days', 'Peak Day']
    hdr_row = tbl_hdr_row + 1
    ws4.row_dimensions[hdr_row].height = 18
    for col, h in enumerate(hdrs, 1):
        c = ws4.cell(row=hdr_row, column=col, value=h)
        c.font = SUBHDR_FONT; c.fill = SUBHDR_FILL
        c.alignment = CENTER; c.border = BORDER

    tbl_data_start = hdr_row + 1
    for i, prod in enumerate(active_cols, tbl_data_start):
        d     = cons_active[prod] if prod in cons_active.columns else pd.Series(dtype=float)
        d_pos = d[d > 0]
        pct   = round(float(d.sum()) / total_all * 100, 1) if total_all > 0 else 0
        area  = short_area(prod)
        peak_date = ''
        if len(d_pos):
            try:
                peak_date = pd.to_datetime(d.idxmax()).strftime('%d-%b-%Y')
            except Exception:
                peak_date = str(d.idxmax())
        row_v = [prod, area,
                 round(float(d.sum()), 1),
                 round(float(d_pos.mean()), 2) if len(d_pos) else 0,
                 round(float(d_pos.max()),  1) if len(d_pos) else 0,
                 f'{pct:.1f}%',
                 int(len(d_pos)),
                 peak_date]
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        ws4.row_dimensions[i].height = 16
        for col, val in enumerate(row_v, 1):
            c = ws4.cell(row=i, column=col, value=val)
            c.font = NORMAL_FONT; c.fill = fill; c.border = BORDER
            c.alignment = LEFT if col <= 2 else CENTER
            if col in (3, 4, 5): c.number_format = '#,##0.0'

    tbl_data_end  = tbl_data_start + len(active_cols) - 1
    grand_row_n   = tbl_data_end + 1
    ws4.row_dimensions[grand_row_n].height = 16
    grand_v = ['TOTAL', '', round(total_all, 1), '', '', '100.0%', '', '']
    for col, val in enumerate(grand_v, 1):
        c = ws4.cell(row=grand_row_n, column=col, value=val)
        c.font = TOTAL_FONT; c.fill = TOTAL_FILL
        c.alignment = CENTER; c.border = BORDER

    # ══════════════════════════════════════════════════════════════════════════
    # HIDDEN HELPER TABLES (col J=10 onward) — chart data sources
    # Label column uses Excel IF formula referencing H2 toggle
    # ══════════════════════════════════════════════════════════════════════════
    HLP_ROW = 3

    # Helper A: Label + Total  (J=10, K=11)
    HA_LBL = 10; HA_VAL = 11
    ws4.cell(row=HLP_ROW, column=HA_LBL, value='Label_A')
    ws4.cell(row=HLP_ROW, column=HA_VAL, value='Total (m\u00b3)')
    for i, prod in enumerate(active_cols, HLP_ROW + 1):
        area  = short_area(prod)
        lw = f'{prod} - {area}' if area else prod
        lo = prod
        ws4.cell(row=i, column=HA_LBL, value=f'=IF($H$2="YES","{lw}","{lo}")')
        ws4.cell(row=i, column=HA_VAL,
                 value=round(float(cons_totals.get(prod, 0)), 1))
    ha_end           = HLP_ROW + len(active_cols)
    totals_label_ref = Reference(ws4, min_col=HA_LBL, min_row=HLP_ROW+1, max_row=ha_end)
    totals_data_ref  = Reference(ws4, min_col=HA_VAL, min_row=HLP_ROW,   max_row=ha_end)

    # Helper B: Avg/Max  (L=12, M=13, N=14)
    HB_LBL = 12; HB_AVG = 13; HB_MAX = 14
    ws4.cell(row=HLP_ROW, column=HB_LBL, value='Label_B')
    ws4.cell(row=HLP_ROW, column=HB_AVG, value='Avg/day')
    ws4.cell(row=HLP_ROW, column=HB_MAX, value='Max/day')
    for i, prod in enumerate(active_cols, HLP_ROW + 1):
        d     = cons_active[prod] if prod in cons_active.columns else pd.Series(dtype=float)
        d_pos = d[d > 0]
        area  = short_area(prod)
        lw = f'{prod} - {area}' if area else prod
        lo = prod
        ws4.cell(row=i, column=HB_LBL, value=f'=IF($H$2="YES","{lw}","{lo}")')
        ws4.cell(row=i, column=HB_AVG,
                 value=round(float(d_pos.mean()), 2) if len(d_pos) else 0)
        ws4.cell(row=i, column=HB_MAX,
                 value=round(float(d_pos.max()),  1) if len(d_pos) else 0)
    hb_end     = HLP_ROW + len(active_cols)
    am_cat_ref = Reference(ws4, min_col=HB_LBL, min_row=HLP_ROW+1, max_row=hb_end)
    am_avg_ref = Reference(ws4, min_col=HB_AVG, min_row=HLP_ROW,   max_row=hb_end)
    am_max_ref = Reference(ws4, min_col=HB_MAX, min_row=HLP_ROW,   max_row=hb_end)

    # Helper C: Cumulative  (O=15, P=16+)
    HC_DATE = 15; HC_START = 16
    ws4.cell(row=HLP_ROW, column=HC_DATE, value='Date')
    for j, prod in enumerate(active_cols):
        area  = short_area(prod)
        lw = f'{prod} - {area}' if area else prod
        lo = prod
        ws4.cell(row=HLP_ROW, column=HC_START+j,
                 value=f'=IF($H$2="YES","{lw}","{lo}")')
    cons_disp2 = cons.iloc[1:]
    for k, (date, _) in enumerate(cons_disp2.iterrows(), HLP_ROW + 1):
        ws4.cell(row=k, column=HC_DATE, value=fmt_date(date))
        for j, prod in enumerate(active_cols):
            if prod in cons_active.columns:
                cumval = float(cons_active[prod].iloc[:k - HLP_ROW].sum())
            else:
                cumval = 0.0
            ws4.cell(row=k, column=HC_START+j, value=round(cumval, 2))
    hc_end       = HLP_ROW + len(cons_disp2)
    cum_date_ref = Reference(ws4, min_col=HC_DATE, min_row=HLP_ROW+1, max_row=hc_end)
    date_ref     = Reference(ws3, min_col=1, min_row=data_start_row, max_row=data_end_row)

    # ══════════════════════════════════════════════════════════════════════════
    # CHARTS — single column, full width, below summary table
    # ══════════════════════════════════════════════════════════════════════════
    cur_row = grand_row_n + 3

    def section_title(txt, row):
        ws4.merge_cells(f'A{row}:H{row}')
        c = ws4.cell(row=row, column=1, value=txt)
        c.font = Font(bold=True, name='Arial', size=11, color='FFFFFF')
        c.fill = SUBHDR_FILL; c.alignment = CENTER
        ws4.row_dimensions[row].height = 20
        return row + 1

    def place(chart, anchor_row):
        chart.width  = CHART_W
        chart.height = CHART_H
        chart.anchor = f'A{anchor_row}'
        ws4.add_chart(chart)
        return anchor_row + rows_per_chart

    # Chart 1: PIE
    cur_row = section_title('Chart 1 \u2014 Consumption Distribution by Unit (Pie)', cur_row)
    pie = PieChart()
    pie.title  = 'Consumption Distribution by Unit'
    pie.style  = 26
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showCatName = True
    pie.dataLabels.showVal     = True
    pie.dataLabels.showPercent = True
    pie.dataLabels.showSerName = False
    pie.add_data(totals_data_ref, titles_from_data=True)
    pie.set_categories(totals_label_ref)
    pie.series[0].title = SeriesLabel(v='Total (m\u00b3)')
    cur_row = place(pie, cur_row)

    # Chart 2: BAR horizontal
    cur_row = section_title('Chart 2 \u2014 Total Water Consumption by Unit (m\u00b3)', cur_row)
    bar_total = BarChart()
    bar_total.type = 'bar'; bar_total.barDir = 'bar'; bar_total.grouping = 'clustered'
    bar_total.title = 'Total Water Consumption by Unit (m\u00b3)'
    bar_total.style = 26
    bar_total.y_axis.title = 'Unit'; bar_total.x_axis.title = 'm\u00b3'
    bar_total.dataLabels = DataLabelList()
    bar_total.dataLabels.showVal = True
    bar_total.add_data(totals_data_ref, titles_from_data=True)
    bar_total.set_categories(totals_label_ref)
    bar_total.series[0].title = SeriesLabel(v='Total (m\u00b3)')
    cur_row = place(bar_total, cur_row)

    # Chart 3: LINE daily trend
    cur_row = section_title('Chart 3 \u2014 Daily Consumption Trend by Unit (m\u00b3/day)', cur_row)
    line_chart = LineChart()
    line_chart.title    = 'Daily Consumption Trend by Unit (m\u00b3/day)'
    line_chart.style    = 10; line_chart.smooth = True; line_chart.grouping = 'standard'
    line_chart.y_axis.title = 'm\u00b3/day'; line_chart.x_axis.title = 'Date'
    for i, prod in enumerate(active_cols):
        dr = Reference(ws3, min_col=i+2, min_row=data_start_row-1, max_row=data_end_row)
        line_chart.add_data(dr, titles_from_data=True)
        area = short_area(prod)
        line_chart.series[i].title = SeriesLabel(v=f'{prod} - {area}' if area else prod)
    line_chart.set_categories(date_ref)
    cur_row = place(line_chart, cur_row)

    # Chart 4: STACKED COL
    cur_row = section_title('Chart 4 \u2014 Daily Consumption Stacked (m\u00b3)', cur_row)
    stacked = BarChart()
    stacked.type = 'col'; stacked.barDir = 'col'; stacked.grouping = 'stacked'
    stacked.title = 'Daily Consumption \u2014 Stacked (m\u00b3)'
    stacked.style = 10
    stacked.y_axis.title = 'm\u00b3/day'; stacked.x_axis.title = 'Date'
    for i, prod in enumerate(active_cols):
        dr = Reference(ws3, min_col=i+2, min_row=data_start_row-1, max_row=data_end_row)
        stacked.add_data(dr, titles_from_data=True)
        area = short_area(prod)
        stacked.series[i].title = SeriesLabel(v=f'{prod} - {area}' if area else prod)
    stacked.set_categories(date_ref)
    cur_row = place(stacked, cur_row)

    # Chart 5: AREA cumulative
    cur_row = section_title('Chart 5 \u2014 Cumulative Consumption (m\u00b3)', cur_row)
    area_chart = AreaChart()
    area_chart.title    = 'Cumulative Consumption (m\u00b3)'
    area_chart.style    = 10; area_chart.grouping = 'standard'
    area_chart.y_axis.title = 'Cumulative m\u00b3'; area_chart.x_axis.title = 'Date'
    for j, prod in enumerate(active_cols):
        dr = Reference(ws4, min_col=HC_START+j, min_row=HLP_ROW, max_row=hc_end)
        area_chart.add_data(dr, titles_from_data=True)
    area_chart.set_categories(cum_date_ref)
    cur_row = place(area_chart, cur_row)

    # Chart 6: CLUSTERED COL avg & max
    cur_row = section_title('Chart 6 \u2014 Avg & Max Daily Consumption by Unit (m\u00b3/day)', cur_row)
    grouped_bar = BarChart()
    grouped_bar.type = 'col'; grouped_bar.barDir = 'col'; grouped_bar.grouping = 'clustered'
    grouped_bar.title = 'Avg & Max Daily Consumption by Unit (m\u00b3/day)'
    grouped_bar.style = 10; grouped_bar.y_axis.title = 'm\u00b3/day'
    grouped_bar.dataLabels = DataLabelList()
    grouped_bar.dataLabels.showVal = True
    grouped_bar.add_data(am_avg_ref, titles_from_data=True)
    grouped_bar.add_data(am_max_ref, titles_from_data=True)
    grouped_bar.set_categories(am_cat_ref)
    grouped_bar.series[0].graphicalProperties.solidFill = '0EA5E9'
    grouped_bar.series[1].graphicalProperties.solidFill = 'EF4444'
    cur_row = place(grouped_bar, cur_row)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ============================================================================
# HTML EXPORT
# ============================================================================

def generate_html_report(
    cons_active: pd.DataFrame,
    cons_totals: pd.Series,
    loc_map: dict,
    selected_prods: list,
    period_str: str,
    total_all: float,
    n_days: int,
) -> str:
    import json
    import plotly.io as pio

    def short_area(prod):
        a = loc_map.get(prod, '')
        return a.replace(f'{prod} - ', '').replace(prod, '').strip(' -')

    avg_daily = float(cons_active.sum(axis=1).mean()) if len(cons_active) > 0 else 0
    peak_val, peak_date_str = get_peak_day_info(cons_active)

    # Build chart JSONs
    fig_pie     = make_pie(cons_totals, loc_map)
    fig_bar     = make_bar_total_html(cons_totals, loc_map)   # HTML version: inside labels
    fig_line    = make_line_daily(cons_active, loc_map, selected_prods)
    fig_stacked = make_stacked_bar(cons_active, loc_map, selected_prods)
    fig_heatmap = make_heatmap(cons_active, selected_prods)
    fig_cumul   = make_cumulative(cons_active, loc_map, selected_prods)
    fig_avgmax  = make_avg_max_bar(cons_active, loc_map, selected_prods)

    def fig_json(fig):
        return pio.to_json(fig)

    # KPI summary table rows
    kpi_rows = ""
    for prod in selected_prods:
        if prod not in cons_active.columns:
            continue
        color = get_unit_color(prod)
        d     = cons_active[prod]
        d_pos = d[d > 0]
        total_v = round(d.sum(), 1)
        avg_v   = round(d_pos.mean(), 2) if len(d_pos) else 0
        max_v   = round(d_pos.max(),  1) if len(d_pos) else 0
        pct_v   = round(total_v / total_all * 100, 1) if total_all > 0 else 0
        area    = short_area(prod)
        kpi_rows += f"""
        <tr>
          <td><span style="display:inline-block;width:10px;height:10px;
              border-radius:50%;background:{color};margin-right:6px;"></span>
              <b>{prod}</b></td>
          <td>{area}</td>
          <td>{total_v:,.1f}</td>
          <td>{avg_v:,.2f}</td>
          <td>{max_v:,.1f}</td>
          <td>{pct_v:.1f}%</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Water Consumption Dashboard</title>
<script src="https://cdn.plot.ly/plotly-2.26.0.min.js"></script>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:Inter,sans-serif;background:#f8fafc;color:#1e293b}}
  .header{{background:white;border-bottom:1px solid #e5e7eb;padding:18px 32px;
            display:flex;align-items:center;gap:16px;}}
  .logo{{width:48px;height:48px;background:linear-gradient(135deg,#0ea5e9,#0284c7);
          border-radius:12px;display:flex;align-items:center;justify-content:center;
          font-size:24px;flex-shrink:0;}}
  .header-title{{font-size:24px;font-weight:800;color:#0c4a6e}}
  .header-sub{{font-size:13px;color:#94a3b8;margin-top:3px}}
  .badge{{background:linear-gradient(90deg,#0ea5e9,#0284c7);color:white;
           padding:8px 18px;border-radius:20px;font-size:13px;font-weight:600;
           display:inline-block;margin:16px 32px 0 auto;}}
  .container{{max-width:1400px;margin:0 auto;padding:24px 32px}}
  .section-title{{font-size:20px;font-weight:700;color:#0c4a6e;margin:28px 0 16px;
                   border-left:4px solid #0ea5e9;padding-left:12px;}}
  .kpi-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:24px}}
  .kpi-card{{background:white;border-radius:12px;padding:20px;
              box-shadow:0 1px 4px rgba(0,0,0,0.08);border-left:4px solid #0ea5e9;}}
  .kpi-label{{font-size:12px;color:#64748b;font-weight:600;text-transform:uppercase;
               letter-spacing:.5px;margin-bottom:6px}}
  .kpi-value{{font-size:28px;font-weight:800;color:#0f172a}}
  .kpi-unit{{font-size:13px;color:#94a3b8;margin-top:2px}}
  .kpi-sub{{font-size:11px;color:#ef4444;font-weight:600;margin-top:4px}}
  .chart-grid{{display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:20px}}
  .chart-full{{margin-bottom:20px}}
  .chart-card{{background:white;border-radius:14px;padding:4px;
                box-shadow:0 1px 6px rgba(0,0,0,0.07);}}
  table{{width:100%;border-collapse:collapse;background:white;border-radius:12px;
         overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,0.08)}}
  thead th{{background:#0c4a6e;color:white;padding:12px 14px;text-align:left;
             font-size:12px;font-weight:700;letter-spacing:.5px;text-transform:uppercase}}
  tbody tr:nth-child(even){{background:#f0f9ff}}
  tbody td{{padding:10px 14px;font-size:13px;border-bottom:1px solid #f1f5f9}}
  .footer{{text-align:center;color:#94a3b8;padding:24px;font-size:12px;margin-top:32px;
            background:white;border-radius:12px}}
  @media(max-width:768px){{
    .kpi-grid{{grid-template-columns:1fr 1fr}}
    .chart-grid{{grid-template-columns:1fr}}
  }}
</style>
</head>
<body>

<div class="header">
  <div class="logo">💧</div>
  <div>
    <div class="header-title">Water Consumption Dashboard</div>
    <div class="header-sub">Daily water consumption analysis from Water Meter Raw Data — Plant 1</div>
  </div>
  <div class="badge">📅 {period_str} &nbsp;|&nbsp; {n_days} days</div>
</div>

<div class="container">

  <div class="section-title">Summary KPIs</div>
  <div class="kpi-grid">
    <div class="kpi-card" style="border-color:#0ea5e9">
      <div class="kpi-label">💧 Total Consumption</div>
      <div class="kpi-value">{total_all:,.1f}</div>
      <div class="kpi-unit">m³</div>
    </div>
    <div class="kpi-card" style="border-color:#10b981">
      <div class="kpi-label">📅 Days of Data</div>
      <div class="kpi-value">{n_days}</div>
      <div class="kpi-unit">days</div>
    </div>
    <div class="kpi-card" style="border-color:#f59e0b">
      <div class="kpi-label">📊 Average / Day</div>
      <div class="kpi-value">{avg_daily:,.1f}</div>
      <div class="kpi-unit">m³/day</div>
    </div>
    <div class="kpi-card" style="border-color:#ef4444">
      <div class="kpi-label">🔺 Peak Day</div>
      <div class="kpi-value">{peak_val:,.1f}</div>
      <div class="kpi-unit">m³</div>
      <div class="kpi-sub">📅 {peak_date_str}</div>
    </div>
  </div>

  <div class="section-title">Unit Statistics</div>
  <table>
    <thead>
      <tr>
        <th>Unit</th><th>Area</th><th>Total (m³)</th>
        <th>Avg/day (m³)</th><th>Max/day (m³)</th><th>% of Total</th>
      </tr>
    </thead>
    <tbody>{kpi_rows}</tbody>
  </table>

  <div class="section-title">Visualisations</div>
  <div class="chart-grid">
    <div class="chart-card"><div id="pie"></div></div>
    <div class="chart-card"><div id="bar"></div></div>
  </div>
  <div class="chart-card chart-full" style="margin-bottom:20px"><div id="cumul"></div></div>
  <div class="chart-card chart-full" style="margin-bottom:20px"><div id="avgmax"></div></div>

  <div class="section-title">Daily Trends</div>
  <div class="chart-card chart-full"><div id="line"></div></div>
  <div class="chart-card chart-full" style="margin-top:20px"><div id="stacked"></div></div>
  <div class="chart-card chart-full" style="margin-top:20px"><div id="heat"></div></div>

  <div class="footer">
    <b>Water Consumption Dashboard</b> &nbsp;|&nbsp;
    PT Güntner Indonesia &nbsp;|&nbsp; Water Monitoring System<br>
    <span style="font-size:11px;">Generated on {pd.Timestamp.now().strftime('%d %B %Y %H:%M')}</span>
  </div>
</div>

<script>
var cfg = {{responsive:true, displayModeBar:true,
            modeBarButtonsToRemove:['lasso2d','select2d']}};

// Store each figure JSON once, then access .data and .layout separately
var _pie     = {fig_json(fig_pie)};
var _bar     = {fig_json(fig_bar)};
var _line    = {fig_json(fig_line)};
var _stacked = {fig_json(fig_stacked)};
var _heat    = {fig_json(fig_heatmap)};
var _cumul   = {fig_json(fig_cumul)};
var _avgmax  = {fig_json(fig_avgmax)};

Plotly.newPlot('pie',     _pie.data,     _pie.layout,     cfg);
Plotly.newPlot('bar',     _bar.data,     _bar.layout,     cfg);
Plotly.newPlot('line',    _line.data,    _line.layout,    cfg);
Plotly.newPlot('stacked', _stacked.data, _stacked.layout, cfg);
Plotly.newPlot('heat',    _heat.data,    _heat.layout,    cfg);
Plotly.newPlot('cumul',   _cumul.data,   _cumul.layout,   cfg);
Plotly.newPlot('avgmax',  _avgmax.data,  _avgmax.layout,  cfg);
</script>
</body>
</html>"""
    return html



st.markdown("""
<div style="background:white;border-bottom:1px solid #e5e7eb;padding:14px 28px;
            display:flex;align-items:center;gap:16px;margin:-4rem -4rem 1.5rem -4rem;">
  <div style="width:48px;height:48px;background:linear-gradient(135deg,#0ea5e9,#0284c7);
              border-radius:12px;display:flex;align-items:center;justify-content:center;
              font-size:24px;flex-shrink:0;">💧</div>
  <div>
    <div style="font-size:26px;font-weight:800;color:#0c4a6e;line-height:1.2;">
      Water Consumption Dashboard</div>
    <div style="font-size:14px;color:#94a3b8;margin-top:2px;">
      Daily water consumption analysis from Water Meter Raw Data — Plant 1</div>
  </div>
</div>
""", unsafe_allow_html=True)


# ============================================================================
# SIDEBAR
# ============================================================================
with st.sidebar:
    st.markdown("### 📁 Upload Data")
    uploaded = st.file_uploader(
        "Upload waterawdata file (*.xlsx)",
        type=['xlsx'],
        help="Upload waterawdata_GI.xlsx or any file with the same format"
    )

    st.markdown("---")
    st.markdown("### ⚙️ Settings")

    if uploaded:
        raw_bytes = uploaded.read()
        with st.spinner("Reading data..."):
            raw_df = load_raw_data(raw_bytes)

        if not raw_df.empty:
            available_prods = sorted(raw_df['Pompa'].unique().tolist())
            default_sel     = [p for p in DEFAULT_PRODS if p in available_prods]
            if not default_sel:
                default_sel = available_prods[:min(7, len(available_prods))]

            selected_prods = st.multiselect(
                "Select Units",
                options=available_prods,
                default=default_sel,
                help="Select one or more units to analyse"
            )

            st.markdown("---")
            st.markdown("### 📅 Date Filter")
            min_date = raw_df['Date'].min().date()
            max_date = raw_df['Date'].max().date()

            date_from = st.date_input("From date", value=min_date,
                                      min_value=min_date)
            date_to   = st.date_input("To date",   value=max_date,
                                      min_value=min_date)

            st.markdown("---")
            st.markdown("### 🧹 Data Preprocessing")
            preprocess_strategy = st.selectbox(
                "Anomaly Handling Strategy",
                ['flag_only', 'clip_to_zero', 'interpolate', 'rolling_median'],
                index=1,
                format_func=lambda x: {
                    'flag_only':      '🔍 Flag Only — no changes, mark anomalies',
                    'clip_to_zero':   '✂️ Clip to Zero — set bad days to prev indicator',
                    'interpolate':    '📈 Interpolate — fill anomalies with time interpolation',
                    'rolling_median': '📊 Rolling Median — replace with 7-day median',
                }[x],
                help="Data asli TIDAK pernah dihapus. Nilai original tersimpan di kolom Water_Indicator_orig."
            )

            st.markdown("---")
            st.markdown("### 🔧 Calculation Options")

            dedup_method = st.selectbox(
                "Deduplication Method",
                ['First', 'Last', 'Max', 'Mean'],
                index=0,
                help=(
                    "When multiple readings exist for the same unit "
                    "on the same day, use this value"
                )
            )

            st.markdown("---")
            st.markdown("### 📊 Display Charts")
            show_pie     = st.checkbox("Pie Chart — Distribution",      value=True)
            show_bar     = st.checkbox("Bar Chart — Total",              value=True)
            show_line    = st.checkbox("Line Chart — Daily Trend",       value=True)
            show_stacked = st.checkbox("Stacked Bar — Daily",            value=True)
            show_heatmap = st.checkbox("Heatmap",                        value=True)
            show_cumul   = st.checkbox("Cumulative Consumption",         value=True)
            show_box     = st.checkbox("Avg & Max Bar Chart",            value=True)
        else:
            selected_prods = []
    else:
        raw_df         = pd.DataFrame()
        selected_prods = DEFAULT_PRODS
        dedup_method   = 'First'
        date_from      = None
        date_to        = None
        show_pie = show_bar = show_line = show_stacked = True
        show_heatmap = show_cumul = show_box = True
        preprocess_strategy = 'clip_to_zero'

    st.markdown("---")
    st.markdown(
        "<div style='font-size:11px;color:#94a3b8;text-align:center;'>"
        "PT Güntner Indonesia<br>Water Monitoring System</div>",
        unsafe_allow_html=True
    )


# ============================================================================
# MAIN CONTENT
# ============================================================================
if uploaded is None or raw_df.empty:
    st.markdown("""
    <div style="text-align:center;padding:60px 20px;">
      <div style="font-size:72px;margin-bottom:20px;">💧</div>
      <div style="font-size:24px;font-weight:700;color:#0c4a6e;margin-bottom:12px;">
        Upload Your Water Meter Data File</div>
      <div style="font-size:15px;color:#64748b;max-width:500px;margin:0 auto 32px;">
        Upload <b>waterawdata_GI.xlsx</b> or a file with the same format via the sidebar
        to start analysing water consumption.
      </div>
      <div style="display:flex;justify-content:center;gap:24px;flex-wrap:wrap;">
        <div style="background:#f0f9ff;border-radius:12px;padding:20px 28px;min-width:160px;">
          <div style="font-size:28px;margin-bottom:8px;">📊</div>
          <div style="font-size:13px;font-weight:600;color:#0c4a6e;">7 Chart Types</div>
        </div>
        <div style="background:#f0fdf4;border-radius:12px;padding:20px 28px;min-width:160px;">
          <div style="font-size:28px;margin-bottom:8px;">⚙️</div>
          <div style="font-size:13px;font-weight:600;color:#14532d;">Flexible Filters</div>
        </div>
        <div style="background:#fff7ed;border-radius:12px;padding:20px 28px;min-width:160px;">
          <div style="font-size:28px;margin-bottom:8px;">📥</div>
          <div style="font-size:13px;font-weight:600;color:#7c2d12;">CSV Export</div>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()


# Apply date filter
if date_from and date_to:
    raw_df = raw_df[(raw_df['Date'].dt.date >= date_from) &
                    (raw_df['Date'].dt.date <= date_to)]

if not selected_prods:
    st.warning("⚠️ Please select at least one unit from the sidebar.")
    st.stop()

with st.spinner("Processing data & detecting anomalies..."):
    result = process_data(raw_df, selected_prods, dedup_method,
                          preprocess_strategy)

if not result:
    st.error("Failed to process data. Please check the file and your unit selection.")
    st.stop()

pivot, cons, dedup_df, loc_map, df_annotated, df_clean = result

cons_active = cons.iloc[1:]
cons_totals = cons_active.sum()
total_all   = cons_totals.sum()

period_str = (f"{fmt_date(cons.index.min())}  →  {fmt_date(cons.index.max())}"
              if len(cons) > 0 else "—")
n_days = len(cons)

# Get peak day info
peak_val, peak_date_str = get_peak_day_info(cons_active)

st.markdown(
    f'<div style="display:flex;justify-content:flex-end;margin-bottom:16px;">'
    f'<div style="background:linear-gradient(90deg,#0ea5e9,#0284c7);color:white;'
    f'padding:10px 22px;border-radius:24px;font-size:14px;font-weight:600;'
    f'box-shadow:0 3px 10px rgba(14,165,233,0.35);">'
    f'📅 {period_str} &nbsp;|&nbsp; {n_days} days of data</div></div>',
    unsafe_allow_html=True
)

# ============================================================================
# TABS
# ============================================================================
tab_ov, tab_daily, tab_table, tab_raw, tab_anom = st.tabs([
    "📊 Overview", "📈 Daily Analysis", "📋 Data Table", "🗃️ Raw Data", "🚨 Anomaly Report"
])


# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 – OVERVIEW
# ─────────────────────────────────────────────────────────────────────────────
with tab_ov:
    st.markdown("## Performance by Unit")

    cols_kpi = st.columns(len(selected_prods))
    for i, prod in enumerate(selected_prods):
        if prod not in cons.columns:
            continue
        area       = loc_map.get(prod, '')
        area_short = area.replace(f'{prod} - ', '').replace(prod, '').strip(' -')
        total_v    = float(cons_totals.get(prod, 0))
        d_vals     = cons_active[prod][cons_active[prod] > 0] if prod in cons_active else pd.Series()
        avg_v      = float(d_vals.mean()) if len(d_vals) > 0 else 0
        max_v      = float(d_vals.max())  if len(d_vals) > 0 else 0
        pct_v      = total_v / total_all * 100 if total_all > 0 else 0
        color      = get_unit_color(prod)
        with cols_kpi[i]:
            st.markdown(
                kpi_card_html(prod, area_short, total_v, avg_v, max_v, pct_v, color),
                unsafe_allow_html=True
            )

    st.markdown("---")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("💧 Total Consumption", f"{total_all:,.1f} m³")
    m2.metric("📅 Days of Data",       f"{n_days} days")
    avg_daily_total = float(cons_active.sum(axis=1).mean()) if len(cons_active) > 0 else 0
    m3.metric("📊 Average / Day",     f"{avg_daily_total:,.1f} m³")
    # FIX: Show peak day with date
    m4.metric("🔺 Peak Day", f"{peak_val:,.1f} m³", delta=peak_date_str)

    st.markdown("---")
    st.markdown("## Visualisations")

    if show_pie or show_bar:
        c1, c2 = st.columns(2)
        if show_pie:
            with c1:
                st.plotly_chart(make_pie(cons_totals, loc_map), use_container_width=True)
        if show_bar:
            with c2:
                st.plotly_chart(make_bar_total(cons_totals, loc_map), use_container_width=True)

    if show_cumul:
        st.plotly_chart(
            make_cumulative(cons_active, loc_map, selected_prods),
            use_container_width=True
        )
    if show_box:
        st.plotly_chart(
            make_avg_max_bar(cons_active, loc_map, selected_prods),
            use_container_width=True
        )

    # ── Export ────────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📥 Export Report")
    st.info(
        "**HTML** — interactive charts, works offline. &nbsp;|&nbsp; "
        "**Excel** — 4 sheets: Summary, Raw Data, Water Usage (m³), Charts."
    )

    exp_col1, exp_col2 = st.columns(2)

    with exp_col1:
        with st.spinner("Building HTML report..."):
            html_report = generate_html_report(
                cons_active, cons_totals, loc_map, selected_prods,
                period_str, total_all, n_days
            )
        fname_html = f"water_report_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.html"
        st.download_button(
            label="⬇️ Download HTML Report (Interactive)",
            data=html_report.encode('utf-8'),
            file_name=fname_html,
            mime='text/html',
            use_container_width=True,
        )

    with exp_col2:
        with st.spinner("Building Excel report..."):
            # Pass the already-filtered raw_df subset for selected units
            raw_for_excel = raw_df[raw_df['Pompa'].isin(selected_prods)].copy()
            excel_bytes = generate_excel_report(
                raw_filtered=raw_for_excel,
                cons=cons,
                cons_active=cons_active,
                cons_totals=cons_totals,
                loc_map=loc_map,
                selected_prods=selected_prods,
                period_str=period_str,
                total_all=total_all,
                n_days=n_days,
            )
        fname_xlsx = f"water_report_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button(
            label="⬇️ Download Excel Report (.xlsx)",
            data=excel_bytes,
            file_name=fname_xlsx,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True,
        )


# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 – DAILY ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────
with tab_daily:
    st.markdown("## Daily Consumption Trends")

    if show_line:
        st.plotly_chart(
            make_line_daily(cons_active, loc_map, selected_prods),
            use_container_width=True
        )

    if show_stacked:
        st.plotly_chart(
            make_stacked_bar(cons_active, loc_map, selected_prods),
            use_container_width=True
        )

    if show_heatmap:
        st.markdown("### Daily Intensity Heatmap")
        st.plotly_chart(make_heatmap(cons_active, selected_prods), use_container_width=True)

    st.markdown("---")
    st.markdown("### Statistics by Unit")
    stats_rows = []
    for prod in selected_prods:
        if prod not in cons_active.columns:
            continue
        d     = cons_active[prod]
        d_pos = d[d > 0]
        stats_rows.append({
            'PROD':         prod,
            'Area':         loc_map.get(prod, '').replace(f'{prod} - ', ''),
            'Total (m³)':   round(d.sum(), 1),
            'Avg/day (m³)': round(d_pos.mean(), 2) if len(d_pos) else 0,
            'Max/day (m³)': round(d_pos.max(),  1) if len(d_pos) else 0,
            'Min/day (m³)': round(d_pos.min(),  1) if len(d_pos) else 0,
            'Std Dev':      round(d_pos.std(),  2) if len(d_pos) else 0,
            '% of Total':   round(d.sum() / total_all * 100, 1) if total_all > 0 else 0,
        })

    stats_df = pd.DataFrame(stats_rows)
    if not stats_df.empty:
        def style_stats(row):
            color = get_unit_color(row['PROD'])
            r, g, b = int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)
            return [f'background-color:rgba({r},{g},{b},0.07)'] * len(row)

        styled = (stats_df.style
                  .apply(style_stats, axis=1)
                  .format({c: '{:,.1f}' for c in stats_df.columns
                           if stats_df[c].dtype in [float, 'float64']})
                  .set_properties(**{'text-align': 'center',
                                     'border': '1px solid #e2e8f0',
                                     'padding': '8px'})
                  .set_table_styles([{
                      'selector': 'thead th',
                      'props': [('background-color', '#0c4a6e'), ('color', 'white'),
                                ('font-weight', 'bold'), ('text-align', 'center'),
                                ('padding', '10px')]
                  }]))
        st.dataframe(styled, use_container_width=True, hide_index=True)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 – DATA TABLE
# ─────────────────────────────────────────────────────────────────────────────
with tab_table:
    st.markdown("## Daily Consumption Table (m³)")
    st.info(
        "Consumption = Today's Water Indicator − Previous Day's Water Indicator. "
        "First row = 0 (no prior data available)."
    )

    disp = cons.copy()
    disp.index = [fmt_date(d) for d in disp.index]
    disp.index.name = 'Date'
    active_cols = [p for p in selected_prods if p in disp.columns]
    disp = disp[active_cols]
    disp['TOTAL'] = disp.sum(axis=1)

    def short_area(unit):
        a = loc_map.get(unit, '')
        return a.replace(f'{unit} - ', '').replace(unit, '').strip(' -')

    loc_row = pd.DataFrame(
        {p: short_area(p) for p in active_cols} | {'TOTAL': ''},
        index=['📍 Location']
    )
    total_row = pd.DataFrame(disp.sum()).T
    total_row.index = ['TOTAL']
    disp_full = pd.concat([loc_row, disp, total_row])

    def highlight_special(row):
        if row.name == 'TOTAL':
            return ['background-color:#0c4a6e;color:white;font-weight:bold'] * len(row)
        if row.name == '📍 Location':
            return ['background-color:#e0f2fe;color:#0c4a6e;font-style:italic;font-weight:600'] * len(row)
        all_idx = list(disp_full.index)
        idx = all_idx.index(row.name)
        return [('background-color:#f0f9ff' if idx % 2 == 0 else 'background-color:white')] * len(row)

    def fmt_cell(v):
        try:
            return f'{float(v):,.1f}'
        except (ValueError, TypeError):
            return str(v)

    styled_tbl = (disp_full.style
                  .apply(highlight_special, axis=1)
                  .format(fmt_cell)
                  .set_properties(**{'text-align': 'center',
                                     'border': '1px solid #e2e8f0',
                                     'padding': '7px'})
                  .set_table_styles([{
                      'selector': 'thead th',
                      'props': [('background-color', '#0284c7'), ('color', 'white'),
                                ('font-weight', 'bold'), ('text-align', 'center'),
                                ('padding', '10px'), ('border', '1px solid #0284c7')]
                  }]))

    st.dataframe(styled_tbl, use_container_width=True, height=520)

    c1, c2 = st.columns(2)
    with c1:
        csv_cons = disp_full.to_csv().encode('utf-8')
        st.download_button(
            "⬇️ Download Daily Consumption Table (CSV)",
            csv_cons, 'water_consumption_daily.csv', 'text/csv',
            use_container_width=True
        )

    st.markdown("---")
    st.markdown("### Monthly Summary")
    cons_monthly_data = []
    tmp   = cons.copy()
    tmp.index = pd.to_datetime([d for d in tmp.index])
    tmp_m = tmp.resample('ME').sum()
    for d, row in tmp_m.iterrows():
        r = {'Month': d.strftime('%B %Y')}
        for prod in selected_prods:
            if prod in row.index:
                r[prod] = round(float(row[prod]), 1)
        r['TOTAL'] = round(sum(row[p] for p in selected_prods if p in row.index), 1)
        cons_monthly_data.append(r)

    if cons_monthly_data:
        monthly_df = pd.DataFrame(cons_monthly_data)
        st.dataframe(monthly_df, use_container_width=True, hide_index=True)

        with c2:
            csv_monthly = monthly_df.to_csv(index=False).encode('utf-8')
            st.download_button(
                "⬇️ Download Monthly Summary (CSV)",
                csv_monthly, 'water_consumption_monthly.csv', 'text/csv',
                use_container_width=True
            )


# ─────────────────────────────────────────────────────────────────────────────
# TAB 4 – RAW DATA
# ─────────────────────────────────────────────────────────────────────────────
with tab_raw:
    st.markdown("## Raw Data (after Deduplication)")
    st.info(
        f"Showing data after deduplication using method: **{dedup_method}**. "
        f"Total rows: {len(dedup_df):,}"
    )

    show_prods_raw = st.multiselect(
        "Filter Units", options=selected_prods,
        default=selected_prods[:3], key='raw_prod_filter'
    )

    raw_view = (dedup_df[dedup_df['Pompa'].isin(show_prods_raw)].copy()
                if show_prods_raw else dedup_df.copy())
    raw_view['Date'] = raw_view['Date'].dt.strftime('%d-%b-%Y')
    raw_view = raw_view.rename(columns={'Water_Indicator': 'Water Indicator (m³)'})
    raw_view['Water Indicator (m³)'] = raw_view['Water Indicator (m³)'].round(2)

    st.dataframe(
        raw_view[['Date', 'Location', 'Pompa', 'Water Indicator (m³)']].reset_index(drop=True),
        use_container_width=True, height=500, hide_index=True
    )

    csv_raw = raw_view.to_csv(index=False).encode('utf-8')
    st.download_button(
        "⬇️ Download Raw Data (CSV)",
        csv_raw, 'water_raw_dedup.csv', 'text/csv'
    )

    st.markdown("---")
    st.markdown("### Pivot Table — Daily Water Indicator")
    pivot_disp = pivot.copy()
    pivot_disp.index = [fmt_date(d) for d in pivot_disp.index]
    pivot_disp = pivot_disp[[p for p in selected_prods if p in pivot_disp.columns]]
    # Tampilkan consumption (daily diff) bukan raw indicator, dan None → 0
    cons_disp_tbl = cons.copy()
    cons_disp_tbl.index = [fmt_date(d) for d in cons_disp_tbl.index]
    cons_disp_tbl = cons_disp_tbl[[p for p in selected_prods if p in cons_disp_tbl.columns]]
    cons_disp_tbl = cons_disp_tbl.fillna(0).round(1)
    # Add TOTAL column
    cons_disp_tbl['TOTAL'] = cons_disp_tbl.sum(axis=1).round(1)
    st.dataframe(cons_disp_tbl, use_container_width=True, height=400)



# ─────────────────────────────────────────────────────────────────────────────
# TAB 5 – ANOMALY REPORT
# ─────────────────────────────────────────────────────────────────────────────
with tab_anom:
    anom_summary = get_anomaly_summary(df_annotated)
    n_anom       = len(anom_summary)
    n_total      = len(df_annotated)
    n_preprocessed = int(df_clean['preprocessed'].sum()) if 'preprocessed' in df_clean.columns else 0

    # ── KPI row ──────────────────────────────────────────────────────────────
    ka, kb, kc, kd = st.columns(4)
    ka.metric("Total Records",      f"{n_total:,}")
    kb.metric("Anomalies Detected", f"{n_anom:,}",
              delta=f"{n_anom/n_total*100:.1f}%" if n_total else None,
              delta_color="inverse")
    kc.metric("Records Corrected",  f"{n_preprocessed:,}",
              help="Rows where Water_Indicator was adjusted by preprocessing")
    kd.metric("Strategy Applied",   preprocess_strategy.replace('_', ' ').title())

    if n_anom == 0:
        st.success("✅ No anomalies detected in selected date range and units.")
    else:
        st.markdown("---")

        # ── Anomaly type breakdown chart ──────────────────────────────────────
        col_chart, col_pie = st.columns([3, 2])
        with col_chart:
            st.markdown("#### Anomaly Count by Type")
            type_counts = (anom_summary['anomaly_type']
                           .value_counts().reset_index()
                           .rename(columns={'index': 'type', 'anomaly_type': 'count',
                                            'count': 'count'}))
            type_counts.columns = ['Type', 'Count']
            color_map = {
                'SPIKE EKSTREM':             '#ef4444',
                'SPIKE TINGGI':              '#f97316',
                'SPIKE':                     '#f59e0b',
                'INPUT ERROR / METER RESET': '#8b5cf6',
                'NILAI NEGATIF':             '#06b6d4',
                'DIGIT PREFIX ERROR':        '#e11d48',
                'PERGANTIAN FLOWMETER':      '#7c3aed',
            }
            fig_bar_a = go.Figure(go.Bar(
                x=type_counts['Count'], y=type_counts['Type'],
                orientation='h',
                marker_color=[color_map.get(t, '#94a3b8') for t in type_counts['Type']],
                text=type_counts['Count'], textposition='outside'
            ))
            fig_bar_a.update_layout(
                height=280, margin=dict(l=0, r=40, t=10, b=10),
                xaxis_title='Count', yaxis_title='',
                plot_bgcolor='white', paper_bgcolor='white',
            )
            st.plotly_chart(fig_bar_a, use_container_width=True)

        with col_pie:
            st.markdown("#### Anomaly Reason Breakdown")
            reason_counts = anom_summary['anomaly_reason'].value_counts().head(8)
            # Shorten labels
            short_reasons = [r[:40] + '…' if len(r) > 40 else r for r in reason_counts.index]
            fig_pie_a = go.Figure(go.Pie(
                labels=short_reasons, values=reason_counts.values,
                hole=0.4, textinfo='percent',
                hovertext=reason_counts.index,
            ))
            fig_pie_a.update_layout(
                height=280, margin=dict(l=0, r=0, t=10, b=10),
                showlegend=True, legend=dict(font_size=9),
                paper_bgcolor='white',
            )
            st.plotly_chart(fig_pie_a, use_container_width=True)

        # ── Anomalies over time ───────────────────────────────────────────────
        st.markdown("#### Anomalies Over Time")
        anom_by_date = (anom_summary.groupby(['Date', 'anomaly_type'])
                        .size().reset_index(name='count'))
        fig_time = go.Figure()
        for atype, color in color_map.items():
            d = anom_by_date[anom_by_date['anomaly_type'] == atype]
            if not d.empty:
                fig_time.add_trace(go.Bar(
                    x=d['Date'], y=d['count'],
                    name=atype, marker_color=color
                ))
        fig_time.update_layout(
            barmode='stack', height=240,
            margin=dict(l=0, r=0, t=10, b=10),
            xaxis_title='Date', yaxis_title='Count',
            plot_bgcolor='white', paper_bgcolor='white',
            legend=dict(orientation='h', y=-0.3, font_size=10),
        )
        st.plotly_chart(fig_time, use_container_width=True)

        # ── Per-pompa anomaly table ───────────────────────────────────────────
        st.markdown("#### Anomalies per Unit (Pompa)")
        per_pompa = (anom_summary.groupby('Pompa')
                     .agg(
                         Total_Anomalies=('anomaly_type', 'count'),
                         Spike_Count=('anomaly_type', lambda x: (x.str.contains('SPIKE')).sum()),
                         Negative_Count=('anomaly_type', lambda x: (x == 'NILAI NEGATIF').sum()),
                         Reset_Count=('anomaly_type', lambda x: (x == 'INPUT ERROR / METER RESET').sum()),
                         Worst_Z=('z_score', lambda x: x.abs().max()),
                     ).reset_index().sort_values('Total_Anomalies', ascending=False))
        per_pompa['Worst_Z'] = per_pompa['Worst_Z'].round(1)
        per_pompa['Status'] = per_pompa['Total_Anomalies'].apply(
            lambda x: '🔴 PERLU AUDIT' if x >= 10 else ('🟡 PERLU CEK' if x >= 3 else '🟢 OK')
        )
        st.dataframe(per_pompa, use_container_width=True, hide_index=True)

        st.markdown("---")

        # ── Detailed anomaly table ────────────────────────────────────────────
        st.markdown("#### Detailed Anomaly List")

        # Filters
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            filter_type = st.multiselect(
                "Filter by Anomaly Type",
                options=anom_summary['anomaly_type'].unique().tolist(),
                default=anom_summary['anomaly_type'].unique().tolist()
            )
        with fc2:
            filter_pompa = st.multiselect(
                "Filter by Unit (Pompa)",
                options=anom_summary['Pompa'].unique().tolist(),
                default=anom_summary['Pompa'].unique().tolist()
            )
        with fc3:
            year_options = sorted(anom_summary['Date'].dt.year.unique().tolist())
            filter_year  = st.multiselect("Filter by Year", options=year_options,
                                           default=year_options)

        disp = anom_summary[
            (anom_summary['anomaly_type'].isin(filter_type)) &
            (anom_summary['Pompa'].isin(filter_pompa)) &
            (anom_summary['Date'].dt.year.isin(filter_year))
        ].copy()
        disp['Date'] = disp['Date'].dt.strftime('%d-%b-%Y')

        st.dataframe(disp, use_container_width=True, height=400, hide_index=True)

        # Download anomaly report CSV
        csv_anom = anom_summary.copy()
        csv_anom['Date'] = csv_anom['Date'].dt.strftime('%Y-%m-%d')
        st.download_button(
            "⬇️ Download Anomaly Report (CSV)",
            csv_anom.to_csv(index=False).encode('utf-8'),
            'anomaly_report.csv', 'text/csv'
        )

        # ── Strategy explanation ──────────────────────────────────────────────
        st.markdown("---")
        st.markdown("#### ℹ️ Preprocessing Strategy Explanation")
        strategy_info = {
            'flag_only': (
                "**🔍 Flag Only** — Semua anomali ditandai tapi **tidak ada nilai yang diubah**. "
                "Data visualisasi menggunakan nilai asli. Cocok untuk audit manual."
            ),
            'clip_to_zero': (
                "**✂️ Clip to Zero** — Untuk hari yang anomali negatif, indikator diganti dengan "
                "nilai hari sebelumnya sehingga usage = 0 (tidak ada konsumsi negatif). "
                "Data asli tetap tersimpan di kolom `Water_Indicator_orig`."
            ),
            'interpolate': (
                "**📈 Interpolate** — Nilai indikator pada hari anomali digantikan dengan "
                "interpolasi linear berdasarkan waktu (titik sebelum dan sesudah). "
                "Data asli tetap tersimpan di kolom `Water_Indicator_orig`."
            ),
            'rolling_median': (
                "**📊 Rolling Median** — Nilai indikator pada hari anomali digantikan dengan "
                "median 7 hari rolling window di sekitar tanggal tersebut. "
                "Data asli tetap tersimpan di kolom `Water_Indicator_orig`."
            ),
        }
        st.info(strategy_info.get(preprocess_strategy, ''))
        st.markdown(
            "> ⚠️ **Integritas Data:** Nilai asli **TIDAK PERNAH dihapus**. "
            "Semua perubahan bersifat computed — file upload Anda tidak berubah. "
            "Kolom `Water_Indicator_orig` selalu menyimpan nilai asli untuk auditability."
        )


st.markdown("---")
st.markdown("""
<div style='text-align:center;color:#94a3b8;padding:1.5rem;
            background:white;border-radius:0.75rem;font-size:13px;'>
  <b style='color:#0c4a6e;'>Water Consumption Dashboard</b> &nbsp;|&nbsp;
  PT Güntner Indonesia &nbsp;|&nbsp; Water Monitoring System
</div>
""", unsafe_allow_html=True)
